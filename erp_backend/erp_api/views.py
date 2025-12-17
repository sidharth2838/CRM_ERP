# =============== IMPORTS ===============
import json
from datetime import datetime, timedelta
from django.utils import timezone

from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.contrib.auth.models import User
from django.db.models import Sum, Count, Q, F, Avg, FloatField

# REST Framework imports
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.decorators import api_view, permission_classes
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework import generics

# Model imports
from .models import *
from .decorators import role_required, admin_required, admin_or_manager_required, finance_required, staff_required
from .serializers import (
    SiteInfoSerializer,
    WebsiteStorySerializer,
    WebsiteTestimonialSerializer,
    WebsiteGallerySerializer,
    WebsiteFAQSerializer,
    WebsitePartnerSerializer,
    WebsiteHeroSectionSerializer,
    WebsiteCollectionsSectionSerializer,
    WebsiteQualitySectionSerializer,
    WebsiteNewsletterSerializer,
    WebsiteFurnitureDetailsSectionSerializer,
    WebsiteTestimonialsSectionSettingsSerializer,
    WebsiteStoriesSectionSettingsSerializer,
)

# =============== AUTHENTICATION VIEWS ===============
class RegisterView(APIView):
    """Register a new user"""
    permission_classes = [AllowAny]
    
    def post(self, request):
        try:
            data = request.data
            
            # Validation
            required_fields = ['username', 'password', 'email']
            for field in required_fields:
                if not data.get(field):
                    return Response({
                        'error': f'Missing required field: {field}'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check if user exists
            if User.objects.filter(username=data['username']).exists():
                return Response({
                    'error': 'Username already exists'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if User.objects.filter(email=data['email']).exists():
                return Response({
                    'error': 'Email already exists'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Create user
            user = User.objects.create_user(
                username=data['username'],
                email=data['email'],
                password=data['password'],
                first_name=data.get('first_name', ''),
                last_name=data.get('last_name', ''),
            )
            
            # Create user profile
            UserProfile.objects.create(
                user=user,
                unique_id=data.get('unique_id', f"USER{user.id:04d}"),
                role=data.get('role', 'customer'),
                phone=data.get('phone', ''),
                department=data.get('department', '')
            )
            
            # Create customer if role is customer
            if data.get('role') == 'customer':
                Customer.objects.create(
                    user=user,
                    customer_code=data.get('customer_code', f"CUST{user.id:04d}"),
                    billing_address=data.get('billing_address', ''),
                    shipping_address=data.get('shipping_address', ''),
                    credit_limit=data.get('credit_limit', 10000),
                    tax_number=data.get('tax_number', '')
                )
            
            # Generate tokens
            refresh = RefreshToken.for_user(user)
            
            return Response({
                'success': True,
                'message': 'User registered successfully',
                'user_id': user.id,
                'tokens': {
                    'refresh': str(refresh),
                    'access': str(refresh.access_token),
                }
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


class LoginView(APIView):
    """Login user and return JWT tokens"""
    permission_classes = [AllowAny]
    
    def post(self, request):
        try:
            data = request.data
            username = data.get('username')
            password = data.get('password')
            
            # Validation
            if not username or not password:
                return Response({
                    'error': 'Username and password are required'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check if user exists
            try:
                user = User.objects.get(username=username)
            except User.DoesNotExist:
                return Response({
                    'error': 'Invalid credentials'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            # Check password
            if not user.check_password(password):
                return Response({
                    'error': 'Invalid credentials'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            # Generate tokens
            refresh = RefreshToken.for_user(user)
            
            # Update last login
            user.last_login = timezone.now()
            user.save()
            
            # Get user profile
            profile = user.userprofile
            
            return Response({
                'success': True,
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'role': profile.role,
                    'unique_id': profile.unique_id
                },
                'tokens': {
                    'refresh': str(refresh),
                    'access': str(refresh.access_token),
                }
            })
            
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# =============== DASHBOARD VIEWS ===============
class DashboardView(APIView):
    """Main dashboard with analytics and metrics"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        try:
            # Get date ranges
            today = datetime.now().date()
            week_ago = today - timedelta(days=7)
            month_ago = today - timedelta(days=30)
            
            # Basic counts
            total_customers = Customer.objects.count()
            total_products = Product.objects.count()
            total_orders = Order.objects.count()
            total_invoices = Invoice.objects.count()
            
            # Today's metrics
            today_orders = Order.objects.filter(order_date__date=today).count()
            today_sales = Order.objects.filter(
                order_date__date=today, 
                status='delivered'
            ).aggregate(total=Sum('grand_total'))['total'] or 0
            
            # This week metrics
            week_orders = Order.objects.filter(order_date__date__gte=week_ago).count()
            week_sales = Order.objects.filter(
                order_date__date__gte=week_ago,
                status='delivered'
            ).aggregate(total=Sum('grand_total'))['total'] or 0
            
            # This month metrics
            month_orders = Order.objects.filter(order_date__date__gte=month_ago).count()
            month_sales = Order.objects.filter(
                order_date__date__gte=month_ago,
                status='delivered'
            ).aggregate(total=Sum('grand_total'))['total'] or 0
            
            # Pending items
            pending_orders = Order.objects.filter(status='pending').count()
            pending_invoices = Invoice.objects.filter(status='sent').count()
            
            # Low stock products
            low_stock = Product.objects.filter(stock_quantity__lt=F('min_stock_level')).count()
            
            # Recent orders (last 10)
            recent_orders = Order.objects.select_related('customer__user').order_by('-order_date')[:10].values(
                'id', 'order_number', 'customer__user__username',
                'status', 'grand_total', 'order_date'
            )
            
            # Top selling products
            top_products = OrderItem.objects.select_related('product').values(
                'product__name', 'product__sku'
            ).annotate(
                total_sold=Sum('quantity'),
                total_revenue=Sum('total_price')
            ).order_by('-total_sold')[:5]
            
            # Sales data for chart (last 7 days)
            sales_data = []
            for i in range(6, -1, -1):
                date = today - timedelta(days=i)
                daily_sales = Order.objects.filter(
                    order_date__date=date,
                    status='delivered'
                ).aggregate(total=Sum('grand_total'))['total'] or 0
                
                sales_data.append({
                    'date': date.strftime('%Y-%m-%d'),
                    'sales': float(daily_sales)
                })
            
            # Orders by status
            orders_by_status = Order.objects.values('status').annotate(
                count=Count('id')
            )
            
            # Revenue by product category
            revenue_by_category = []
            categories = ProductCategory.objects.all()
            for category in categories:
                category_revenue = OrderItem.objects.filter(
                    product__category=category
                ).aggregate(total=Sum('total_price'))['total'] or 0
                
                if category_revenue > 0:
                    revenue_by_category.append({
                        'category': category.name,
                        'revenue': float(category_revenue)
                    })
            
            # Recent customers
            recent_customers = Customer.objects.select_related('user').order_by('-created_at')[:5].values(
                'id', 'customer_code', 'user__username', 'user__email', 'created_at'
            )
            
            # Activity log
            recent_activities = ActivityLog.objects.select_related('user').order_by('-created_at')[:10].values(
                'user__username', 'action', 'table_name', 'created_at'
            )
            
            data = {
                'stats': {
                    'total_customers': total_customers,
                    'total_products': total_products,
                    'total_orders': total_orders,
                    'total_invoices': total_invoices,
                    'today_orders': today_orders,
                    'today_sales': float(today_sales),
                    'week_orders': week_orders,
                    'week_sales': float(week_sales),
                    'month_orders': month_orders,
                    'month_sales': float(month_sales),
                    'pending_orders': pending_orders,
                    'pending_invoices': pending_invoices,
                    'low_stock': low_stock,
                },
                'charts': {
                    'sales_data': sales_data,
                    'orders_by_status': list(orders_by_status),
                    'revenue_by_category': revenue_by_category,
                },
                'recent': {
                    'orders': list(recent_orders),
                    'customers': list(recent_customers),
                    'activities': list(recent_activities),
                },
                'top_products': list(top_products),
            }
            
            return Response(data)
            
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# =============== CUSTOMER API VIEWS ===============
class CustomersAPIView(APIView):
    """List and create customers"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        """Get paginated list of customers with filters"""
        try:
            # Get query parameters
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 10))
            search = request.GET.get('search', '').strip()
            customer_type_filter = request.GET.get('customer_type', '').strip()
            
            # Build query
            queryset = Customer.objects.select_related('user').all()
            
            # Apply filters
            if search:
                queryset = queryset.filter(
                    Q(user__first_name__icontains=search) |
                    Q(user__last_name__icontains=search) |
                    Q(user__email__icontains=search) |
                    Q(customer_code__icontains=search)
                )
            
            if customer_type_filter:
                queryset = queryset.filter(customer_type=customer_type_filter)
            
            # Get total count
            total_count = queryset.count()
            
            # Apply pagination
            start = (page - 1) * page_size
            end = start + page_size
            paginated_queryset = queryset[start:end]
            
            # Format results
            results = []
            for customer in paginated_queryset:
                user = customer.user
                profile = getattr(user, 'userprofile', None)
                
                customer_data = {
                    'id': customer.id,
                    'customer_code': customer.customer_code,
                    'first_name': user.first_name or '',
                    'last_name': user.last_name or '',
                    'email': user.email or '',
                    'phone': profile.phone if profile else '',
                    'company': customer.company.name if customer.company else '',
                    'customer_type': getattr(customer, 'customer_type', 'retail'),
                    'credit_limit': float(customer.credit_limit or 0),
                    'balance': float(customer.balance or 0),
                    'tax_number': customer.tax_number or '',
                    'billing_address': customer.billing_address or '',
                    'shipping_address': customer.shipping_address or '',
                    'created_at': customer.created_at.isoformat() if customer.created_at else ''
                }
                results.append(customer_data)
            
            # Get stats
            stats = {
                'total_customers': Customer.objects.count(),
                'total_balance': float(Customer.objects.aggregate(Sum('balance'))['balance__sum'] or 0),
                'new_this_month': Customer.objects.filter(
                    created_at__month=datetime.now().month,
                    created_at__year=datetime.now().year
                ).count()
            }
            
            return Response({
                'success': True,
                'results': results,
                'count': total_count,
                'stats': stats,
                'page': page,
                'page_size': page_size,
                'total_pages': (total_count + page_size - 1) // page_size
            })
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def post(self, request):
        """Create new customer"""
        try:
            data = request.data
            
            # Generate username if not provided
            if not data.get('username'):
                email = data.get('email', '')
                if email:
                    username = email.split('@')[0]
                else:
                    username = f"cust{Customer.objects.count() + 1:04d}"
                data['username'] = username
            
            # Check if user exists
            if User.objects.filter(username=data['username']).exists():
                return Response({
                    'success': False,
                    'error': 'Username already exists'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if User.objects.filter(email=data.get('email', '')).exists():
                return Response({
                    'success': False,
                    'error': 'Email already exists'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Create user
            user = User.objects.create_user(
                username=data['username'],
                email=data.get('email', ''),
                password=data.get('password', 'DefaultPass@123'),
                first_name=data.get('first_name', ''),
                last_name=data.get('last_name', ''),
            )
            
            # Update user profile if it was auto-created by signal
            # The signal should have already created one with a generated unique_id
            if hasattr(user, 'userprofile'):
                profile = user.userprofile
                # Update phone and other fields if provided
                if 'phone' in data:
                    profile.phone = data['phone']
                profile.save()
            else:
                # Create user profile if it doesn't exist (shouldn't happen with signal)
                # Generate unique_id ensuring it doesn't conflict
                unique_id = data.get('unique_id', f"CUST{user.id:04d}")
                
                # Ensure unique_id is actually unique
                counter = 1
                original_unique_id = unique_id
                while UserProfile.objects.filter(unique_id=unique_id).exists():
                    unique_id = f"{original_unique_id}_{counter}"
                    counter += 1
                
                profile = UserProfile.objects.create(
                    user=user,
                    unique_id=unique_id,
                    role='customer',
                    phone=data.get('phone', ''),
                    department='Customer'
                )
            
            # Generate customer code if not provided
            customer_code = data.get('customer_code', f"CUST{user.id:04d}")
            
            # Handle company - either by ID or by name lookup
            company = None
            company_data = data.get('company')
            if company_data:
                try:
                    # Try treating it as an ID first
                    if isinstance(company_data, int) or (isinstance(company_data, str) and company_data.isdigit()):
                        company = Company.objects.get(id=int(company_data))
                    else:
                        # Try exact match first
                        try:
                            company = Company.objects.get(name__exact=company_data)
                        except Company.DoesNotExist:
                            # Try case-insensitive match, get first result
                            company = Company.objects.filter(name__icontains=company_data).first()
                except Company.DoesNotExist:
                    company = None
            
            # Create customer
            customer = Customer.objects.create(
                user=user,
                customer_code=customer_code,
                company=company,
                customer_type=data.get('customer_type', 'retail'),
                credit_limit=float(data.get('credit_limit', 10000)),
                tax_number=data.get('tax_number', ''),
                billing_address=data.get('billing_address', ''),
                shipping_address=data.get('shipping_address', ''),
                balance=0.0
            )
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='CREATE_CUSTOMER',
                table_name='customers',
                record_id=customer.id,
                new_values={'customer_code': customer.customer_code}
            )
            
            return Response({
                'success': True,
                'message': 'Customer created successfully',
                'customer_id': customer.id,
                'customer_code': customer.customer_code
            })
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


class CustomerDetailAPIView(APIView):
    """Retrieve, update, or delete a customer"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request, customer_id):
        """Get customer details"""
        try:
            customer = Customer.objects.select_related('user').get(id=customer_id)
            user = customer.user
            profile = getattr(user, 'userprofile', None)
            
            customer_data = {
                'id': customer.id,
                'customer_code': customer.customer_code,
                'first_name': user.first_name or '',
                'last_name': user.last_name or '',
                'email': user.email or '',
                'phone': profile.phone if profile else '',
                'company': customer.company.name if customer.company else '',
                'customer_type': getattr(customer, 'customer_type', 'retail'),
                'credit_limit': float(customer.credit_limit or 0),
                'balance': float(customer.balance or 0),
                'tax_number': customer.tax_number or '',
                'billing_address': customer.billing_address or '',
                'shipping_address': customer.shipping_address or '',
                'created_at': customer.created_at.isoformat() if customer.created_at else ''
            }
            
            return Response({
                'success': True,
                'customer': customer_data
            })
            
        except Customer.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Customer not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, customer_id):
        """Update customer - Finance, Admin, Manager only"""
        try:
            # Check authorization - only finance, admin, manager can update
            if request.user.is_authenticated:
                try:
                    user_profile = UserProfile.objects.get(user=request.user)
                    if user_profile.role not in ['finance', 'admin', 'manager']:
                        return Response({
                            'success': False,
                            'error': 'Access Denied: Only Finance, Admin, or Manager can update customers'
                        }, status=status.HTTP_403_FORBIDDEN)
                except UserProfile.DoesNotExist:
                    return Response({
                        'success': False,
                        'error': 'User profile not found'
                    }, status=status.HTTP_403_FORBIDDEN)
            else:
                return Response({
                    'success': False,
                    'error': 'Authentication required'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            customer = Customer.objects.get(id=customer_id)
            user = customer.user
            
            data = request.data
            
            # Update user fields
            if 'first_name' in data:
                user.first_name = data['first_name']
            if 'last_name' in data:
                user.last_name = data['last_name']
            if 'email' in data:
                user.email = data['email']
            user.save()
            
            # Update user profile
            if hasattr(user, 'userprofile'):
                profile = user.userprofile
                if 'phone' in data:
                    profile.phone = data['phone']
                profile.save()
            
            # Update customer fields
            update_fields = [
                'company', 'customer_type', 'credit_limit',
                'tax_number', 'billing_address', 'shipping_address'
            ]
            
            for field in update_fields:
                if field in data:
                    if field == 'company':
                        # Handle company lookup
                        company_data = data[field]
                        if company_data:
                            try:
                                if isinstance(company_data, int) or (isinstance(company_data, str) and company_data.isdigit()):
                                    company = Company.objects.get(id=int(company_data))
                                else:
                                    # Try exact match first
                                    try:
                                        company = Company.objects.get(name__exact=company_data)
                                    except Company.DoesNotExist:
                                        # Try case-insensitive match, get first result
                                        company = Company.objects.filter(name__icontains=company_data).first()
                                setattr(customer, field, company)
                            except Company.DoesNotExist:
                                setattr(customer, field, None)
                        else:
                            setattr(customer, field, None)
                    else:
                        setattr(customer, field, data[field])
            
            customer.save()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='UPDATE_CUSTOMER',
                table_name='customers',
                record_id=customer.id,
                new_values=data
            )
            
            return Response({
                'success': True,
                'message': 'Customer updated successfully'
            })
            
        except Customer.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Customer not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, customer_id):
        """Delete customer - Any authenticated user"""
        try:
            if not request.user.is_authenticated:
                return Response({
                    'success': False,
                    'error': 'Authentication required'
                }, status=status.HTTP_401_UNAUTHORIZED)
            customer = Customer.objects.get(id=customer_id)
            user = customer.user
            customer_code = customer.customer_code
            customer.delete()
            user.delete()
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='DELETE_CUSTOMER',
                table_name='customers',
                record_id=customer_id,
                old_values={'customer_code': customer_code}
            )
            return Response({
                'success': True,
                'message': 'Customer deleted successfully'
            })
            
        except Customer.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Customer not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


class CustomerExportAPIView(APIView):
    """Export customers to CSV"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        try:
            customers = Customer.objects.select_related('user').all()
            
            # Format data for CSV
            csv_data = []
            headers = [
                'Customer Code', 'First Name', 'Last Name', 'Email', 'Phone',
                'Company', 'Type', 'Status', 'Credit Limit', 'Balance',
                'Tax Number', 'Billing Address', 'Shipping Address', 'Notes',
                'Created Date'
            ]
            
            for customer in customers:
                user = customer.user
                profile = getattr(user, 'userprofile', None)
                
                row = [
                    customer.customer_code,
                    user.first_name or '',
                    user.last_name or '',
                    user.email or '',
                    profile.phone if profile else '',
                    getattr(customer, 'company', ''),
                    getattr(customer, 'customer_type', 'retail'),
                    getattr(customer, 'status', 'active'),
                    str(customer.credit_limit),
                    str(customer.balance),
                    customer.tax_number or '',
                    customer.billing_address or '',
                    customer.shipping_address or '',
                    getattr(customer, 'notes', ''),
                    customer.created_at.strftime('%Y-%m-%d %H:%M:%S') if customer.created_at else ''
                ]
                csv_data.append(row)
            
            return Response({
                'success': True,
                'headers': headers,
                'data': csv_data
            })
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# =============== COMPANIES API VIEWS ===============
class CompaniesAPIView(APIView):
    """List and create companies"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        """Get paginated list of companies with filters"""
        try:
            # Get query parameters
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 10))
            search = request.GET.get('search', '').strip()
            
            # Build query
            queryset = Company.objects.select_related('created_by').all()
            
            # Apply filters
            if search:
                queryset = queryset.filter(
                    Q(name__icontains=search) |
                    Q(address__icontains=search) |
                    Q(email__icontains=search) |
                    Q(phone__icontains=search)
                )
            
            # Get total count
            total_count = queryset.count()
            
            # Apply pagination
            start = (page - 1) * page_size
            end = start + page_size
            paginated_queryset = queryset[start:end]
            
            # Format results
            results = []
            for company in paginated_queryset:
                company_data = {
                    'id': company.id,
                    'name': company.name,
                    'address': company.address or '',
                    'phone': company.phone or '',
                    'email': company.email or '',
                    'contact_person': company.contact_person or '',
                    'created_at': company.created_at.isoformat() if company.created_at else '',
                    'created_by': company.created_by.username if company.created_by else ''
                }
                results.append(company_data)
            
            # Get stats
            stats = {
                'total_companies': Company.objects.count(),
            }
            
            return Response({
                'success': True,
                'results': results,
                'count': total_count,
                'stats': stats,
                'page': page,
                'page_size': page_size,
                'total_pages': (total_count + page_size - 1) // page_size
            })
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def post(self, request):
        """Create new company"""
        try:
            data = request.data
            
            # Validate required fields
            required_fields = ['name']
            for field in required_fields:
                if not data.get(field):
                    return Response({
                        'success': False,
                        'error': f'Missing required field: {field}'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check if company with same name already exists
            if Company.objects.filter(name=data['name']).exists():
                return Response({
                    'success': False,
                    'error': 'Company with this name already exists'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Create company
            company = Company.objects.create(
                name=data['name'],
                address=data.get('address', ''),
                phone=data.get('phone', ''),
                email=data.get('email', ''),
                contact_person=data.get('contact_person', ''),
                created_by=request.user if request.user.is_authenticated else None
            )
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='CREATE_COMPANY',
                table_name='companies',
                record_id=company.id,
                new_values=data
            )
            
            return Response({
                'success': True,
                'message': 'Company created successfully',
                'company': {
                    'id': company.id,
                    'name': company.name,
                    'address': company.address or '',
                    'phone': company.phone or '',
                    'email': company.email or '',
                    'contact_person': company.contact_person or '',
                    'created_at': company.created_at.isoformat() if company.created_at else '',
                    'created_by': company.created_by.username if company.created_by else ''
                }
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


class CompanyDetailAPIView(APIView):
    """Retrieve, update, or delete a company"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request, company_id):
        """Get company details"""
        try:
            company = Company.objects.select_related('created_by').get(id=company_id)
            
            company_data = {
                'id': company.id,
                'name': company.name,
                'address': company.address or '',
                'phone': company.phone or '',
                'email': company.email or '',
                'contact_person': company.contact_person or '',
                'created_at': company.created_at.isoformat() if company.created_at else '',
                'created_by': company.created_by.username if company.created_by else ''
            }
            
            return Response({
                'success': True,
                'company': company_data
            })
            
        except Company.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Company not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, company_id):
        """Update company"""
        try:
            if not request.user.is_authenticated:
                return Response({
                    'success': False,
                    'error': 'Authentication required'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            company = Company.objects.get(id=company_id)
            data = request.data
            
            # Update fields if provided
            update_fields = ['name', 'address', 'phone', 'email', 'contact_person']
            
            for field in update_fields:
                if field in data:
                    setattr(company, field, data[field])
            
            company.save()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='UPDATE_COMPANY',
                table_name='companies',
                record_id=company.id,
                new_values=data
            )
            
            return Response({
                'success': True,
                'message': 'Company updated successfully'
            })
            
        except Company.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Company not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def patch(self, request, company_id):
        """Partially update company"""
        try:
            if not request.user.is_authenticated:
                return Response({
                    'success': False,
                    'error': 'Authentication required'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            company = Company.objects.get(id=company_id)
            data = request.data
            
            # Update fields if provided
            update_fields = ['name', 'address', 'phone', 'email', 'contact_person']
            
            for field in update_fields:
                if field in data:
                    setattr(company, field, data[field])
            
            company.save()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='UPDATE_COMPANY',
                table_name='companies',
                record_id=company.id,
                new_values=data
            )
            
            return Response({
                'success': True,
                'message': 'Company updated successfully'
            })
            
        except Company.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Company not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, company_id):
        """Delete company"""
        try:
            if not request.user.is_authenticated:
                return Response({
                    'success': False,
                    'error': 'Authentication required'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            company = Company.objects.get(id=company_id)
            company_name = company.name
            company.delete()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='DELETE_COMPANY',
                table_name='companies',
                record_id=company_id,
                old_values={'name': company_name}
            )
            
            return Response({
                'success': True,
                'message': 'Company deleted successfully'
            })
            
        except Company.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Company not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# =============== PRODUCT API VIEWS ===============
class ProductsAPIView(APIView):
    """List and create products"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        """Get paginated list of products with filters"""
        try:
            # Get query parameters
            search = request.GET.get('search', '').strip()
            category = request.GET.get('category', '').strip()
            status_filter = request.GET.get('status', '').strip()
            
            # Build query
            queryset = Product.objects.select_related('category').all()
            
            # Apply filters
            if search:
                queryset = queryset.filter(
                    Q(name__icontains=search) |
                    Q(sku__icontains=search) |
                    Q(description__icontains=search)
                )
            
            if category:
                queryset = queryset.filter(category__name=category)
            
            if status_filter == 'active':
                queryset = queryset.filter(is_active=True)
            elif status_filter == 'inactive':
                queryset = queryset.filter(is_active=False)
            
            products = queryset.values(
                'id', 'name', 'sku', 'category__name',
                'price', 'cost', 'stock_quantity', 'min_stock_level', 
                'is_active', 'description', 'created_at', 'image'
            )
            
            product_list = []
            for p in products:
                image_url = None
                if p['image']:
                    image_url = request.build_absolute_uri(f"/media/{p['image']}")
                
                product_list.append({
                    'id': p['id'],
                    'name': p['name'],
                    'sku': p['sku'],
                    'category': p['category__name'] or 'Uncategorized',
                    'price': float(p['price']),
                    'cost': float(p['cost']) if p['cost'] else float(p['price']) * 0.7,
                    'stock_quantity': p['stock_quantity'],
                    'min_stock_level': p['min_stock_level'],
                    'is_active': p['is_active'],
                    'description': p['description'] or '',
                    'created_at': p['created_at'].isoformat() if p['created_at'] else '',
                    'image_url': image_url
                })
            
            # Calculate stats
            total_products = Product.objects.count()
            active_products = Product.objects.filter(is_active=True).count()
            total_value = Product.objects.aggregate(
                total=Sum(F('price') * F('stock_quantity'))
            )['total'] or 0
            
            low_stock = Product.objects.filter(
                stock_quantity__gt=0,
                stock_quantity__lte=F('min_stock_level')
            ).count()
            
            stats = {
                'total_products': total_products,
                'active_products': active_products,
                'total_value': float(total_value),
                'low_stock': low_stock
            }
            
            return Response({
                'success': True,
                'results': product_list,
                'count': len(product_list),
                'stats': stats
            })
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def post(self, request):
        """Create new product"""
        try:
            data = request.data
            
            # Validate required fields
            required_fields = ['name', 'sku', 'price']
            for field in required_fields:
                if field not in data or not str(data[field]).strip():
                    return Response({
                        'success': False, 
                        'error': f'Missing required field: {field}'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check if SKU already exists
            if Product.objects.filter(sku=data['sku']).exists():
                return Response({
                    'success': False, 
                    'error': 'Product with this SKU already exists'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get or create category
            category = None
            if data.get('category'):
                category, _ = ProductCategory.objects.get_or_create(
                    name=data['category']
                )
            
            # Create product
            product = Product.objects.create(
                name=data['name'].strip(),
                sku=data['sku'].strip(),
                description=data.get('description', '').strip(),
                category=category,
                price=float(data['price']),
                cost=float(data.get('cost', 0)),
                stock_quantity=int(data.get('stock_quantity', 0)),
                min_stock_level=int(data.get('min_stock_level', 10)),
                is_active=bool(data.get('is_active', True)),
                created_by=request.user if request.user.is_authenticated else None
            )
            
            # Handle image upload
            if 'image' in request.FILES:
                product.image = request.FILES['image']
                product.save()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='CREATE_PRODUCT',
                table_name='products',
                record_id=product.id,
                new_values={'sku': product.sku, 'name': product.name}
            )
            
            return Response({
                'success': True,
                'message': 'Product created successfully',
                'product_id': product.id,
                'product': {
                    'id': product.id,
                    'name': product.name,
                    'sku': product.sku,
                    'category': product.category.name if product.category else 'Uncategorized'
                }
            })
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


class ProductDetailAPIView(APIView):
    """Retrieve, update, or delete a product"""
    permission_classes = [AllowAny]
    
    def get(self, request, product_id):
        """Get product details"""
        try:
            product = Product.objects.select_related('category').get(id=product_id)
            
            image_url = None
            if product.image:
                image_url = request.build_absolute_uri(f"/media/{product.image}")
            
            product_data = {
                'id': product.id,
                'name': product.name,
                'sku': product.sku,
                'category': product.category.name if product.category else 'Uncategorized',
                'price': float(product.price),
                'cost': float(product.cost) if product.cost else float(product.price) * 0.7,
                'stock_quantity': product.stock_quantity,
                'min_stock_level': product.min_stock_level,
                'is_active': product.is_active,
                'description': product.description or '',
                'created_at': product.created_at.isoformat() if product.created_at else '',
                'created_by': product.created_by.username if product.created_by else 'System',
                'image_url': image_url
            }
            
            return Response({
                'success': True,
                'product': product_data
            })
            
        except Product.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Product not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, product_id):
        """Update product - Finance, Admin, Manager only"""
        try:
            # Check authorization - only finance, admin, manager can update
            if request.user.is_authenticated:
                try:
                    user_profile = UserProfile.objects.get(user=request.user)
                    if user_profile.role not in ['finance', 'admin', 'manager']:
                        return Response({
                            'success': False,
                            'error': 'Access Denied: Only Finance, Admin, or Manager can update products'
                        }, status=status.HTTP_403_FORBIDDEN)
                except UserProfile.DoesNotExist:
                    return Response({
                        'success': False,
                        'error': 'User profile not found'
                    }, status=status.HTTP_403_FORBIDDEN)
            else:
                return Response({
                    'success': False,
                    'error': 'Authentication required'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            product = Product.objects.get(id=product_id)
            data = request.data
            
            # Update category if provided
            if 'category' in data and data['category']:
                category, _ = ProductCategory.objects.get_or_create(
                    name=data['category']
                )
                product.category = category
            
            # Update other fields
            update_fields = [
                'name', 'sku', 'description', 'price', 'cost',
                'stock_quantity', 'min_stock_level', 'is_active'
            ]
            
            for field in update_fields:
                if field in data:
                    if field in ['price', 'cost']:
                        setattr(product, field, float(data[field]))
                    elif field in ['stock_quantity', 'min_stock_level']:
                        setattr(product, field, int(data[field]))
                    elif field == 'is_active':
                        setattr(product, field, bool(data[field]))
                    else:
                        setattr(product, field, data[field])
            
            # Handle image upload
            if 'image' in request.FILES:
                product.image = request.FILES['image']
            
            product.save()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='UPDATE_PRODUCT',
                table_name='products',
                record_id=product.id,
                new_values={k: v for k, v in data.items() if k in update_fields}
            )
            
            return Response({
                'success': True,
                'message': 'Product updated successfully'
            })
            
        except Product.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Product not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, product_id):
        """Delete product - Finance, Admin, Manager only"""
        try:
            # Check authorization - only finance, admin, manager can delete
            if request.user.is_authenticated:
                try:
                    user_profile = UserProfile.objects.get(user=request.user)
                    if user_profile.role not in ['finance', 'admin', 'manager']:
                        return Response({
                            'success': False,
                            'error': 'Access Denied: Only Finance, Admin, or Manager can delete products'
                        }, status=status.HTTP_403_FORBIDDEN)
                except UserProfile.DoesNotExist:
                    return Response({
                        'success': False,
                        'error': 'User profile not found'
                    }, status=status.HTTP_403_FORBIDDEN)
            else:
                return Response({
                    'success': False,
                    'error': 'Authentication required'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            product = Product.objects.get(id=product_id)
            
            # Log before deletion
            product_name = product.name
            product_sku = product.sku
            
            # Delete product
            product.delete()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='DELETE_PRODUCT',
                table_name='products',
                record_id=product_id,
                old_values={'name': product_name, 'sku': product_sku}
            )
            
            return Response({
                'success': True,
                'message': 'Product deleted successfully'
            })
            
        except Product.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Product not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


class ProductCategoriesAPIView(APIView):
    """Get list of product categories"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        try:
            categories = ProductCategory.objects.filter(is_active=True).values('id', 'name')
            
            return Response({
                'success': True,
                'categories': list(categories)
            })
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# =============== ORDER API VIEWS ===============
class OrdersAPIView(APIView):
    """List and create orders"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        """Get paginated list of orders with filters"""
        try:
            # Get query parameters
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 10))
            search = request.GET.get('search', '').strip()
            status_filter = request.GET.get('status', '').strip()
            payment_filter = request.GET.get('payment_status', '').strip()
            
            # Build query
            queryset = Order.objects.select_related('customer__user').all()
            
            # Apply filters
            if search:
                queryset = queryset.filter(
                    Q(order_number__icontains=search) |
                    Q(customer__user__first_name__icontains=search) |
                    Q(customer__user__last_name__icontains=search) |
                    Q(customer__user__email__icontains=search)
                )
            
            if status_filter:
                queryset = queryset.filter(status=status_filter)
            
            if payment_filter:
                queryset = queryset.filter(payment_status=payment_filter)
            
            # Get total count
            total_count = queryset.count()
            
            # Apply pagination
            start = (page - 1) * page_size
            end = start + page_size
            paginated_queryset = queryset[start:end]
            
            # Format results
            results = []
            for order in paginated_queryset:
                user = order.customer.user if order.customer and order.customer.user else None
                
                results.append({
                    'id': order.id,
                    'order_number': order.order_number or f"ORD{order.id:05d}",
                    'customer_name': f"{user.first_name if user else ''} {user.last_name if user else ''}".strip() or 'Unknown Customer',
                    'customer_email': user.email if user else '',
                    'status': order.status,
                    'payment_status': order.payment_status,
                    'grand_total': float(order.grand_total),
                    'total_amount': float(order.total_amount),
                    'tax_amount': float(order.tax_amount),
                    'discount_amount': float(order.discount_amount),
                    'order_date': order.order_date.isoformat() if order.order_date else '',
                    'shipping_address': order.shipping_address or ''
                })
            
            # Calculate stats
            stats = {
                'total_orders': Order.objects.count(),
                'pending_orders': Order.objects.filter(status='pending').count(),
                'processing_orders': Order.objects.filter(status='processing').count(),
                'today_revenue': Order.objects.filter(
                    order_date__date=datetime.now().date(),
                    status='delivered'
                ).aggregate(total=Sum('grand_total'))['total'] or 0
            }
            
            return Response({
                'success': True,
                'results': results,
                'count': total_count,
                'stats': stats,
                'page': page,
                'page_size': page_size,
                'total_pages': (total_count + page_size - 1) // page_size
            })
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def post(self, request):
        """Create new order"""
        try:
            data = request.data
            
            # Validate required fields
            if 'customer_id' not in data:
                return Response({
                    'success': False,
                    'error': 'Customer ID is required'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get customer
            try:
                customer = Customer.objects.get(id=data['customer_id'])
            except Customer.DoesNotExist:
                return Response({
                    'success': False,
                    'error': 'Customer not found'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Generate order number
            order_number = data.get('order_number', f"ORD{Order.objects.count() + 1:05d}")
            
            # Calculate totals
            total_amount = float(data.get('total_amount', 0))
            tax_amount = float(data.get('tax_amount', 0))
            discount_amount = float(data.get('discount_amount', 0))
            grand_total = total_amount + tax_amount - discount_amount
            
            # Create order
            order = Order.objects.create(
                order_number=order_number,
                customer=customer,
                total_amount=total_amount,
                tax_amount=tax_amount,
                discount_amount=discount_amount,
                grand_total=grand_total,
                status=data.get('status', 'pending'),
                payment_status=data.get('payment_status', 'pending'),
                shipping_address=data.get('shipping_address', customer.shipping_address or ''),
                notes=data.get('notes', ''),
                created_by=request.user if request.user.is_authenticated else None
            )
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='CREATE_ORDER',
                table_name='orders',
                record_id=order.id,
                new_values={'order_number': order.order_number}
            )
            
            return Response({
                'success': True,
                'message': 'Order created successfully',
                'order_id': order.id,
                'order_number': order.order_number
            })
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


class OrderDetailAPIView(APIView):
    """Retrieve, update, or delete an order"""
    permission_classes = [AllowAny]
    
    def get(self, request, order_id):
        """Get order details"""
        try:
            order = Order.objects.select_related('customer__user').get(id=order_id)
            
            order_data = {
                'id': order.id,
                'order_number': order.order_number,
                'customer_id': order.customer.id if order.customer else None,
                'customer_name': f"{order.customer.user.first_name if order.customer and order.customer.user else ''} {order.customer.user.last_name if order.customer and order.customer.user else ''}".strip() or 'Unknown Customer',
                'customer_email': order.customer.user.email if order.customer and order.customer.user else '',
                'customer_phone': order.customer.user.userprofile.phone if order.customer and order.customer.user and hasattr(order.customer.user, 'userprofile') else '',
                'status': order.status,
                'payment_status': order.payment_status,
                'total_amount': float(order.total_amount),
                'tax_amount': float(order.tax_amount),
                'discount_amount': float(order.discount_amount),
                'grand_total': float(order.grand_total),
                'shipping_address': order.shipping_address or '',
                'notes': order.notes or '',
                'order_date': order.order_date.isoformat() if order.order_date else ''
            }
            
            return Response({
                'success': True,
                'order': order_data
            })
            
        except Order.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Order not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, order_id):
        """Update order"""
        try:
            order = Order.objects.get(id=order_id)
            data = request.data
            
            # Update order fields
            update_fields = [
                'status', 'payment_status',
                'shipping_address', 'notes'
            ]
            
            for field in update_fields:
                if field in data:
                    setattr(order, field, data[field])
            
            # Update numeric fields
            if 'total_amount' in data:
                order.total_amount = float(data['total_amount'])
            if 'tax_amount' in data:
                order.tax_amount = float(data['tax_amount'])
            if 'discount_amount' in data:
                order.discount_amount = float(data['discount_amount'])
            
            # Recalculate grand total
            order.grand_total = order.total_amount + order.tax_amount - order.discount_amount
            
            order.save()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='UPDATE_ORDER',
                table_name='orders',
                record_id=order.id,
                new_values=data
            )
            
            return Response({
                'success': True,
                'message': 'Order updated successfully'
            })
            
        except Order.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Order not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, order_id):
        """Delete order"""
        try:
            order = Order.objects.get(id=order_id)
            
            # Log before deletion
            order_number = order.order_number
            
            # Delete order
            order.delete()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='DELETE_ORDER',
                table_name='orders',
                record_id=order_id,
                old_values={'order_number': order_number}
            )
            
            return Response({
                'success': True,
                'message': 'Order deleted successfully'
            })
            
        except Order.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Order not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# =============== INVOICES API ===============
class InvoicesAPIView(APIView):
    """List and create invoices"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        """Get paginated list of invoices with filters"""
        try:
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 10))
            search = request.GET.get('search', '').strip()
            status_filter = request.GET.get('status', '').strip()
            
            queryset = Invoice.objects.select_related('customer__user').all()
            
            if search:
                queryset = queryset.filter(
                    Q(invoice_number__icontains=search) |
                    Q(customer__user__first_name__icontains=search) |
                    Q(customer__user__last_name__icontains=search)
                )
            
            if status_filter:
                queryset = queryset.filter(status=status_filter)
            
            total_count = queryset.count()
            start = (page - 1) * page_size
            end = start + page_size
            paginated_queryset = queryset[start:end]
            
            results = []
            for invoice in paginated_queryset:
                user = invoice.customer.user if invoice.customer and invoice.customer.user else None
                results.append({
                    'id': invoice.id,
                    'invoice_number': invoice.invoice_number or f"INV{invoice.id:05d}",
                    'customer_name': f"{user.first_name if user else ''} {user.last_name if user else ''}".strip() or 'Unknown',
                    'customer_email': user.email if user else '',
                    'status': invoice.status,
                    'amount': float(invoice.total_amount),
                    'paid_amount': float(invoice.paid_amount or 0),
                    'outstanding': float((invoice.total_amount or 0) - (invoice.paid_amount or 0)),
                    'issue_date': invoice.invoice_date.isoformat() if invoice.invoice_date else '',
                    'due_date': invoice.due_date.isoformat() if invoice.due_date else ''
                })
            
            stats = {
                'total_invoices': Invoice.objects.count(),
                'paid_invoices': Invoice.objects.filter(status='paid').count(),
                'outstanding_amount': float(Invoice.objects.aggregate(
                    total=Sum(F('total_amount') - F('paid_amount'), output_field=FloatField())
                )['total'] or 0),
                'total_revenue': float(Invoice.objects.aggregate(Sum('total_amount'))['total_amount__sum'] or 0)
            }
            
            return Response({
                'success': True,
                'results': results,
                'count': total_count,
                'stats': stats,
                'page': page,
                'page_size': page_size,
                'total_pages': (total_count + page_size - 1) // page_size
            })
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def post(self, request):
        """Create new invoice"""
        try:
            data = request.data
            
            if 'customer_id' not in data:
                return Response({
                    'success': False,
                    'error': 'Customer ID is required'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                customer = Customer.objects.get(id=data['customer_id'])
            except Customer.DoesNotExist:
                return Response({
                    'success': False,
                    'error': 'Customer not found'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            invoice_number = data.get('invoice_number', f"INV{Invoice.objects.count() + 1:05d}")
            amount = float(data.get('amount', 0))
            
            invoice = Invoice.objects.create(
                invoice_number=invoice_number,
                customer=customer,
                total_amount=amount,
                paid_amount=float(data.get('paid_amount', 0)),
                status=data.get('status', 'draft'),
                invoice_date=data.get('invoice_date') or datetime.now().date(),
                due_date=data.get('due_date'),
                notes=data.get('notes', ''),
                created_by=request.user if request.user.is_authenticated else None
            )
            
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='CREATE_INVOICE',
                table_name='invoices',
                record_id=invoice.id,
                new_values={'invoice_number': invoice.invoice_number}
            )
            
            return Response({
                'success': True,
                'message': 'Invoice created successfully',
                'invoice_id': invoice.id,
                'invoice_number': invoice.invoice_number
            })
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


class InvoiceDetailAPIView(APIView):
    """Retrieve, update, or delete an invoice"""
    permission_classes = [AllowAny]
    
    def get(self, request, invoice_id):
        """Get invoice details"""
        try:
            invoice = Invoice.objects.select_related('customer__user').get(id=invoice_id)
            
            invoice_data = {
                'id': invoice.id,
                'invoice_number': invoice.invoice_number,
                'customer_id': invoice.customer.id if invoice.customer else None,
                'customer_name': f"{invoice.customer.user.first_name if invoice.customer and invoice.customer.user else ''} {invoice.customer.user.last_name if invoice.customer and invoice.customer.user else ''}".strip() or 'Unknown',
                'status': invoice.status,
                'amount': float(invoice.total_amount or 0),
                'paid_amount': float(invoice.paid_amount or 0),
                'outstanding': float((invoice.total_amount or 0) - (invoice.paid_amount or 0)),
                'issue_date': invoice.invoice_date.isoformat() if invoice.invoice_date else '',
                'due_date': invoice.due_date.isoformat() if invoice.due_date else '',
                'notes': invoice.notes or '',
                'created_at': invoice.created_at.isoformat() if invoice.created_at else ''
            }
            
            return Response({
                'success': True,
                'invoice': invoice_data
            })
        except Invoice.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Invoice not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, invoice_id):
        """Update invoice"""
        try:
            invoice = Invoice.objects.get(id=invoice_id)
            data = request.data
            
            old_status = invoice.status
            
            invoice.status = data.get('status', invoice.status)
            invoice.total_amount = float(data.get('amount', invoice.total_amount))
            invoice.paid_amount = float(data.get('paid_amount', invoice.paid_amount or 0))
            invoice.due_date = data.get('due_date', invoice.due_date)
            invoice.notes = data.get('notes', invoice.notes)
            invoice.save()
            
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='UPDATE_INVOICE',
                table_name='invoices',
                record_id=invoice.id,
                old_values={'status': old_status},
                new_values={'status': invoice.status}
            )
            
            return Response({
                'success': True,
                'message': 'Invoice updated successfully'
            })
        except Invoice.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Invoice not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, invoice_id):
        """Delete invoice"""
        try:
            invoice = Invoice.objects.get(id=invoice_id)
            invoice_number = invoice.invoice_number
            invoice.delete()
            
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='DELETE_INVOICE',
                table_name='invoices',
                record_id=invoice_id,
                old_values={'invoice_number': invoice_number}
            )
            
            return Response({
                'success': True,
                'message': 'Invoice deleted successfully'
            })
        except Invoice.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Invoice not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# =============== LEADS API ===============
class LeadsAPIView(APIView):
    """List and create leads"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        """Get paginated list of leads with filters"""
        try:
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 10))
            search = request.GET.get('search', '').strip()
            status_filter = request.GET.get('status', '').strip()
            
            queryset = Lead.objects.all()
            
            if search:
                queryset = queryset.filter(
                    Q(lead_number__icontains=search) |
                    Q(contact_person__icontains=search) |
                    Q(company_name__icontains=search) |
                    Q(email__icontains=search) |
                    Q(phone__icontains=search)
                )
            
            if status_filter:
                queryset = queryset.filter(status=status_filter)
            
            total_count = queryset.count()
            start = (page - 1) * page_size
            end = start + page_size
            paginated_queryset = queryset[start:end]
            
            results = []
            for lead in paginated_queryset:
                results.append({
                    'id': lead.id,
                    'lead_number': lead.lead_number or f"LEAD{lead.id:04d}",
                    'contact_person': lead.contact_person or '',
                    'company_name': lead.company_name or '',
                    'email': lead.email or '',
                    'phone': lead.phone or '',
                    'source': lead.source or '',
                    'status': lead.status or 'new',
                    'estimated_value': float(lead.estimated_value or 0),
                    'notes': lead.notes or '',
                    'created_at': lead.created_at.isoformat() if lead.created_at else ''
                })
            
            stats = {
                'total_leads': Lead.objects.count(),
                'new_leads': Lead.objects.filter(status='new').count(),
                'qualified_leads': Lead.objects.filter(status='qualified').count(),
                'converted_leads': Lead.objects.filter(status='converted').count()
            }
            
            return Response({
                'success': True,
                'results': results,
                'count': total_count,
                'stats': stats,
                'page': page,
                'page_size': page_size,
                'total_pages': (total_count + page_size - 1) // page_size
            })
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def post(self, request):
        """Create new lead"""
        try:
            data = request.data
            
            lead_number = data.get('lead_number', f"LEAD{Lead.objects.count() + 1:04d}")
            
            lead = Lead.objects.create(
                lead_number=lead_number,
                contact_person=data.get('contact_person', ''),
                company_name=data.get('company_name', ''),
                email=data.get('email', ''),
                phone=data.get('phone', ''),
                source=data.get('source', 'other'),
                status=data.get('status', 'new'),
                estimated_value=float(data.get('estimated_value', 0)) if data.get('estimated_value') else None,
                notes=data.get('notes', ''),
                created_by=request.user if request.user.is_authenticated else None
            )
            
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='CREATE_LEAD',
                table_name='leads',
                record_id=lead.id,
                new_values={'lead_number': lead.lead_number}
            )
            
            return Response({
                'success': True,
                'message': 'Lead created successfully',
                'lead_id': lead.id,
                'lead_number': lead.lead_number
            })
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


class LeadDetailAPIView(APIView):
    """Retrieve, update, or delete a lead"""
    permission_classes = [AllowAny]
    
    def get(self, request, lead_id):
        """Get lead details"""
        try:
            lead = Lead.objects.get(id=lead_id)
            
            lead_data = {
                'id': lead.id,
                'lead_number': lead.lead_number,
                'contact_person': lead.contact_person or '',
                'company_name': lead.company_name or '',
                'email': lead.email or '',
                'phone': lead.phone or '',
                'source': lead.source or '',
                'status': lead.status or 'new',
                'estimated_value': float(lead.estimated_value or 0),
                'notes': lead.notes or '',
                'created_at': lead.created_at.isoformat() if lead.created_at else ''
            }
            
            return Response({
                'success': True,
                'lead': lead_data
            })
        except Lead.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Lead not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, lead_id):
        """Update lead"""
        try:
            lead = Lead.objects.get(id=lead_id)
            data = request.data
            
            old_status = lead.status
            
            lead.contact_person = data.get('contact_person', lead.contact_person)
            lead.company_name = data.get('company_name', lead.company_name)
            lead.email = data.get('email', lead.email)
            lead.phone = data.get('phone', lead.phone)
            lead.source = data.get('source', lead.source)
            lead.status = data.get('status', lead.status)
            if 'estimated_value' in data and data['estimated_value']:
                lead.estimated_value = float(data['estimated_value'])
            lead.notes = data.get('notes', lead.notes)
            lead.save()
            
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='UPDATE_LEAD',
                table_name='leads',
                record_id=lead.id,
                old_values={'status': old_status},
                new_values={'status': lead.status}
            )
            
            return Response({
                'success': True,
                'message': 'Lead updated successfully'
            })
        except Lead.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Lead not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, lead_id):
        """Delete lead"""
        try:
            lead = Lead.objects.get(id=lead_id)
            lead_number = lead.lead_number
            lead.delete()
            
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='DELETE_LEAD',
                table_name='leads',
                record_id=lead_id,
                old_values={'lead_number': lead_number}
            )
            
            return Response({
                'success': True,
                'message': 'Lead deleted successfully'
            })
        except Lead.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Lead not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# =============== LEADS EXPORT/IMPORT ===============
class LeadsExportAPIView(APIView):
    """Export leads to Excel"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Export all leads to Excel file"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Leads"
            
            # Define header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Add headers
            headers = ['Lead #', 'Contact Person', 'Company', 'Email', 'Phone', 'Source', 'Status', 'Estimated Value', 'Notes', 'Created At']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add data
            leads = Lead.objects.all().order_by('-created_at')
            for row_idx, lead in enumerate(leads, 2):
                ws.cell(row=row_idx, column=1).value = lead.lead_number or f"LEAD{lead.id:04d}"
                ws.cell(row=row_idx, column=2).value = lead.contact_person or ''
                ws.cell(row=row_idx, column=3).value = lead.company_name or ''
                ws.cell(row=row_idx, column=4).value = lead.email or ''
                ws.cell(row=row_idx, column=5).value = lead.phone or ''
                ws.cell(row=row_idx, column=6).value = lead.source or ''
                ws.cell(row=row_idx, column=7).value = lead.status or ''
                ws.cell(row=row_idx, column=8).value = float(lead.estimated_value or 0)
                ws.cell(row=row_idx, column=9).value = lead.notes or ''
                ws.cell(row=row_idx, column=10).value = lead.created_at.strftime('%Y-%m-%d %H:%M:%S') if lead.created_at else ''
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 30
            ws.column_dimensions['J'].width = 18
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="leads_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            wb.save(response)
            
            return response
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


class LeadsImportAPIView(APIView):
    """Import leads from Excel"""
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        """Import leads from uploaded Excel file"""
        try:
            from openpyxl import load_workbook
            from io import BytesIO
            
            if 'file' not in request.FILES:
                return Response({
                    'success': False,
                    'error': 'No file provided'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            uploaded_file = request.FILES['file']
            
            # Load workbook
            wb = load_workbook(BytesIO(uploaded_file.read()))
            ws = wb.active
            
            imported_count = 0
            error_count = 0
            errors = []
            
            # Skip header row
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    if not row[0]:  # Skip empty rows
                        continue
                    
                    # Extract data from row
                    lead_number = row[0] or f"LEAD{Lead.objects.count() + imported_count + 1:04d}"
                    contact_person = row[1] or ''
                    company_name = row[2] or ''
                    email = row[3] or ''
                    phone = row[4] or ''
                    source = row[5] or 'other'
                    status = row[6] or 'new'
                    estimated_value = float(row[7]) if row[7] else None
                    notes = row[8] or ''
                    
                    # Create or update lead
                    lead, created = Lead.objects.update_or_create(
                        lead_number=lead_number,
                        defaults={
                            'contact_person': contact_person,
                            'company_name': company_name,
                            'email': email,
                            'phone': phone,
                            'source': source,
                            'status': status,
                            'estimated_value': estimated_value,
                            'notes': notes,
                            'created_by': request.user if request.user.is_authenticated else None
                        }
                    )
                    
                    imported_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f"Row {row_idx}: {str(e)}")
                    continue
            
            # Log the import
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='IMPORT_LEADS',
                table_name='leads',
                new_values={'imported_count': imported_count, 'error_count': error_count}
            )
            
            return Response({
                'success': True,
                'imported': imported_count,
                'errors': error_count,
                'error_details': errors if errors else None,
                'message': f'Import completed: {imported_count} leads imported, {error_count} errors'
            })
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# =============== PAYMENTS API ===============
class PaymentsAPIView(APIView):
    """List and create payments"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        """Get paginated list of payments with filters"""
        try:
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 10))
            search = request.GET.get('search', '').strip()
            payment_method_filter = request.GET.get('payment_method', '').strip()
            
            queryset = Payment.objects.select_related('invoice__customer__user').all()
            
            if search:
                queryset = queryset.filter(
                    Q(payment_number__icontains=search) |
                    Q(invoice__invoice_number__icontains=search)
                )
            
            if payment_method_filter:
                queryset = queryset.filter(payment_method=payment_method_filter)
            
            total_count = queryset.count()
            start = (page - 1) * page_size
            end = start + page_size
            paginated_queryset = queryset[start:end]
            
            results = []
            for payment in paginated_queryset:
                invoice = payment.invoice
                customer = invoice.customer if invoice else None
                user = customer.user if customer and customer.user else None
                
                results.append({
                    'id': payment.id,
                    'payment_number': payment.payment_number or f"PAY{payment.id:05d}",
                    'invoice_number': invoice.invoice_number if invoice else '',
                    'customer_name': f"{user.first_name if user else ''} {user.last_name if user else ''}".strip() or 'Unknown',
                    'amount': float(payment.amount or 0),
                    'payment_method': payment.payment_method or '',
                    'payment_date': payment.payment_date.isoformat() if payment.payment_date else '',
                    'created_at': payment.created_at.isoformat() if payment.created_at else ''
                })
            
            stats = {
                'total_payments': Payment.objects.count(),
                'total_amount': float(Payment.objects.aggregate(Sum('amount'))['amount__sum'] or 0),
                'credit_card_payments': Payment.objects.filter(payment_method='credit_card').count()
            }
            
            return Response({
                'success': True,
                'results': results,
                'count': total_count,
                'stats': stats,
                'page': page,
                'page_size': page_size,
                'total_pages': (total_count + page_size - 1) // page_size
            })
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def post(self, request):
        """Create new payment"""
        try:
            data = request.data
            
            if 'invoice_id' not in data:
                return Response({
                    'success': False,
                    'error': 'Invoice ID is required'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                invoice = Invoice.objects.get(id=data['invoice_id'])
            except Invoice.DoesNotExist:
                return Response({
                    'success': False,
                    'error': 'Invoice not found'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            payment_number = data.get('payment_number', f"PAY{Payment.objects.count() + 1:05d}")
            amount = float(data.get('amount', 0))
            
            payment = Payment.objects.create(
                payment_number=payment_number,
                invoice=invoice,
                amount=amount,
                payment_method=data.get('payment_method', ''),
                status=data.get('status', 'pending'),
                payment_date=data.get('payment_date') or datetime.now().date(),
                reference_number=data.get('reference_number', ''),
                notes=data.get('notes', ''),
                created_by=request.user if request.user.is_authenticated else None
            )
            
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='CREATE_PAYMENT',
                table_name='payments',
                record_id=payment.id,
                new_values={'payment_number': payment.payment_number}
            )
            
            return Response({
                'success': True,
                'message': 'Payment created successfully',
                'payment_id': payment.id,
                'payment_number': payment.payment_number
            })
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


class PaymentDetailAPIView(APIView):
    """Retrieve, update, or delete a payment"""
    permission_classes = [AllowAny]
    
    def get(self, request, payment_id):
        """Get payment details"""
        try:
            payment = Payment.objects.select_related('invoice__customer__user').get(id=payment_id)
            
            invoice = payment.invoice
            customer = invoice.customer if invoice else None
            user = customer.user if customer and customer.user else None
            
            payment_data = {
                'id': payment.id,
                'payment_number': payment.payment_number,
                'invoice_id': invoice.id if invoice else None,
                'invoice_number': invoice.invoice_number if invoice else '',
                'customer_name': f"{user.first_name if user else ''} {user.last_name if user else ''}".strip() or 'Unknown',
                'amount': float(payment.amount or 0),
                'payment_method': payment.payment_method or '',
                'reference_number': payment.reference_number or '',
                'notes': payment.notes or '',
                'payment_date': payment.payment_date.isoformat() if payment.payment_date else '',
                'created_at': payment.created_at.isoformat() if payment.created_at else ''
            }
            
            return Response({
                'success': True,
                'payment': payment_data
            })
        except Payment.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Payment not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, payment_id):
        """Update payment"""
        try:
            payment = Payment.objects.get(id=payment_id)
            data = request.data
            
            old_status = payment.status
            
            payment.amount = float(data.get('amount', payment.amount))
            payment.payment_method = data.get('payment_method', payment.payment_method)
            payment.status = data.get('status', payment.status)
            payment.reference_number = data.get('reference_number', payment.reference_number)
            payment.notes = data.get('notes', payment.notes)
            payment.payment_date = data.get('payment_date', payment.payment_date)
            payment.save()
            
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='UPDATE_PAYMENT',
                table_name='payments',
                record_id=payment.id,
                old_values={'status': old_status},
                new_values={'status': payment.status}
            )
            
            return Response({
                'success': True,
                'message': 'Payment updated successfully'
            })
        except Payment.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Payment not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, payment_id):
        """Delete payment"""
        try:
            payment = Payment.objects.get(id=payment_id)
            payment_number = payment.payment_number
            payment.delete()
            
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='DELETE_PAYMENT',
                table_name='payments',
                record_id=payment_id,
                old_values={'payment_number': payment_number}
            )
            
            return Response({
                'success': True,
                'message': 'Payment deleted successfully'
            })
        except Payment.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Payment not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# =============== REPORTS VIEW ===============
class ReportsView(APIView):
    """Generate reports"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        report_type = request.GET.get('type', 'sales')
        
        if report_type == 'sales':
            # Last 30 days sales report
            end_date = datetime.now().date()
            start_date = end_date - timedelta(days=30)
            
            sales_report = Order.objects.filter(
                order_date__date__range=[start_date, end_date],
                status='delivered'
            ).extra({
                'date': "DATE(order_date)"
            }).values('date').annotate(
                total_sales=Sum('grand_total'),
                order_count=Count('id')
            ).order_by('date')
            
            return Response({
                'report_type': 'sales',
                'period': f'{start_date} to {end_date}',
                'data': list(sales_report)
            })
        
        elif report_type == 'inventory':
            inventory_report = Product.objects.values(
                'category__name'
            ).annotate(
                total_products=Count('id'),
                total_value=Sum(F('price') * F('stock_quantity')),

                low_stock=Count('id', filter=Q(stock_quantity__lt=F('min_stock_level')))
            )
            
            return Response({
                'report_type': 'inventory',
                'data': list(inventory_report)
            })
        
        elif report_type == 'customers':
            # Customer statistics
            customer_report = Customer.objects.values(
                'customer_type'
            ).annotate(
                total=Count('id'),
                avg_credit_limit=Avg('credit_limit'),
                total_balance=Sum('balance')
            )
            
            return Response({
                'report_type': 'customers',
                'data': list(customer_report)
            })
        
        return Response({
            'error': 'Invalid report type'
        }, status=status.HTTP_400_BAD_REQUEST)


# =============== LEGACY FUNCTIONS ===============
@csrf_exempt
def api_add_lead(request):
    """Add new lead"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Generate lead number if not provided
            lead_number = data.get('lead_number', f"LEAD{Lead.objects.count() + 1:04d}")
            
            lead = Lead.objects.create(
                lead_number=lead_number,
                company_name=data.get('company_name', ''),
                contact_person=data.get('contact_person', ''),
                email=data.get('email', ''),
                phone=data.get('phone', ''),
                source=data.get('source', 'website'),
                status=data.get('status', 'new'),
                estimated_value=float(data.get('estimated_value', 0)),
                notes=data.get('notes', ''),
                created_by=request.user if request.user.is_authenticated else None
            )
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='CREATE_LEAD',
                table_name='leads',
                record_id=lead.id,
                new_values={'lead_number': lead.lead_number}
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Lead added successfully',
                'lead_id': lead.id,
                'lead_number': lead.lead_number
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)


@csrf_exempt
def api_get_leads(request):
    """Get all leads"""
    try:
        leads = Lead.objects.all().values(
            'id', 'lead_number', 'company_name', 'contact_person',
            'email', 'phone', 'source', 'status', 'estimated_value', 'created_at'
        ).order_by('-created_at')
        lead_list = []
        for l in leads:
            lead_list.append({
                'id': l['id'],
                'lead_number': l['lead_number'],
                'company_name': l['company_name'],
                'contact_person': l['contact_person'],
                'email': l['email'],
                'phone': l['phone'],
                'source': l['source'],
                'status': l['status'],
                'estimated_value': float(l['estimated_value']) if l['estimated_value'] else None,
                'created_at': l['created_at'].isoformat() if l['created_at'] else ''
            })
        return JsonResponse({'results': lead_list})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def api_update_order_status(request, order_id):
    """Update order status"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            order = Order.objects.get(id=order_id)
            
            # Log old values
            old_status = order.status;
            
            # Update status
            order.status = data['status']
            order.save()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='UPDATE_ORDER_STATUS',
                table_name='orders',
                record_id=order.id,
                old_values={'status': old_status},
                new_values={'status': order.status}
            )
            
            return JsonResponse({'success': True, 'message': 'Order status updated'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    
    return JsonResponse({'success': False, 'message': 'Invalid method'}, status=405)


@csrf_exempt
def api_search(request):
    """Search across all models"""
    query = request.GET.get('q', '')
    
    try:
        results = {
            'customers': list(Customer.objects.filter(
                Q(user__username__icontains=query) |
                Q(customer_code__icontains=query) |
                Q(user__email__icontains=query)
            ).values('id', 'customer_code', 'user__username')[:5]),
            
            'products': list(Product.objects.filter(
                Q(name__icontains=query) |
                Q(sku__icontains=query) |
                Q(description__icontains=query)
            ).values('id', 'name', 'sku', 'price')[:5]),
            
            'orders': list(Order.objects.filter(
                Q(order_number__icontains=query) |
                Q(customer__user__username__icontains=query)
            ).values('id', 'order_number', 'customer__user__username', 'grand_total')[:5]),
            
            'leads': list(Lead.objects.filter(
                Q(company_name__icontains=query) |
                Q(contact_person__icontains=query) |
                Q(email__icontains=query)
            ).values('id', 'lead_number', 'company_name', 'contact_person')[:5]),
        }
        
        return JsonResponse(results)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def api_customer_detail_legacy(request, customer_id):
    """Legacy: Get, update, or delete a specific customer"""
    try:
        customer = Customer.objects.select_related('user').get(id=customer_id)
        user = customer.user
        
        if request.method == 'GET':
            # Return customer details
            return JsonResponse({
                'success': True,
                'customer': {
                    'id': customer.id,
                    'customer_code': customer.customer_code,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'email': user.email,
                    'phone': getattr(customer.user.userprofile, 'phone', ''),
                    'company': getattr(customer, 'company', ''),
                    'customer_type': getattr(customer, 'customer_type', 'retail'),
                    'status': getattr(customer, 'status', 'active'),
                    'credit_limit': float(customer.credit_limit),
                    'balance': float(customer.balance),
                    'tax_number': customer.tax_number,
                    'billing_address': customer.billing_address,
                    'shipping_address': customer.shipping_address,
                    'notes': getattr(customer, 'notes', ''),
                    'created_at': customer.created_at.isoformat(),
                    'updated_at': customer.updated_at.isoformat()
                }
            })
        
        elif request.method == 'PUT':
            # Update customer
            data = json.loads(request.body)
            
            # Update user fields
            user.first_name = data.get('first_name', user.first_name)
            user.last_name = data.get('last_name', user.last_name)
            user.email = data.get('email', user.email)
            user.save()
            
            # Update phone in profile
            if hasattr(user, 'userprofile'):
                profile = user.userprofile
                profile.phone = data.get('phone', profile.phone)
                profile.save()
            
            # Update customer fields
            customer.company = data.get('company', getattr(customer, 'company', ''))
            customer.customer_type = data.get('customer_type', getattr(customer, 'customer_type', 'retail'))
            customer.status = data.get('status', getattr(customer, 'status', 'active'))
            customer.credit_limit = data.get('credit_limit', customer.credit_limit)
            customer.tax_number = data.get('tax_number', customer.tax_number)
            customer.billing_address = data.get('billing_address', customer.billing_address)
            customer.shipping_address = data.get('shipping_address', customer.shipping_address)
            customer.notes = data.get('notes', getattr(customer, 'notes', ''))
            customer.save()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='UPDATE_CUSTOMER',
                table_name='customers',
                record_id=customer.id,
                new_values=data
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Customer updated successfully'
            })
        
        elif request.method == 'DELETE':
            # Delete customer
            customer_code = customer.customer_code
            user_id = user.id
            customer.delete()
            user.delete()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='DELETE_CUSTOMER',
                table_name='customers',
                record_id=customer_id,
                old_values={'customer_code': customer_code}
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Customer deleted successfully'
            })
    
    except Customer.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Customer not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)


@csrf_exempt
def api_customers_list_legacy(request):
    """Legacy: Get all customers with filtering and pagination"""
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    
    try:
        # Get query parameters
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 10))
        search = request.GET.get('search', '').strip()
        status_filter = request.GET.get('status', '').strip()
        customer_type_filter = request.GET.get('customer_type', '').strip()
        
        # Build query
        queryset = Customer.objects.select_related('user').all()
        
        # Apply search filter
        if search:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search) |
                Q(user__email__icontains=search) |
                Q(customer_code__icontains=search)
            )
        
        # Apply status filter
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Apply customer type filter
        if customer_type_filter:
            queryset = queryset.filter(customer_type=customer_type_filter)
        
        # Get total count
        total_count = queryset.count()
        
        # Get stats
        stats = {
            'total_customers': Customer.objects.count(),
            'active_customers': Customer.objects.filter(status='active').count(),
            'total_balance': float(Customer.objects.aggregate(Sum('balance'))['balance__sum'] or 0),
            'new_this_month': Customer.objects.filter(
                created_at__month=datetime.now().month,
                created_at__year=datetime.now().year
            ).count()
        }
        
        # Apply pagination
        start = (page - 1) * page_size
        end = start + page_size
        queryset = queryset[start:end]
        
        # Format results
        results = []
        for customer in queryset:
            user = customer.user
            profile = getattr(user, 'userprofile', None)
            
            results.append({
                'id': customer.id,
                'customer_code': customer.customer_code,
                'first_name': user.first_name or '',
                'last_name': user.last_name or '',
                'email': user.email or '',
                'phone': profile.phone if profile else '',
                'company': getattr(customer, 'company', ''),
                'customer_type': getattr(customer, 'customer_type', 'retail'),
                'status': getattr(customer, 'status', 'active'),
                'credit_limit': float(customer.credit_limit or 0),
                'balance': float(customer.balance or 0),
                'created_at': customer.created_at.isoformat() if customer.created_at else '',
                'updated_at': customer.updated_at.isoformat() if customer.updated_at else ''
            })
        
        return JsonResponse({
            'success': True,
            'results': results,
            'count': total_count,
            'stats': stats
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@csrf_exempt
def api_delete_customer_legacy(request, customer_id):
    """Legacy: Delete a customer by ID - Finance, Admin, Manager only"""
    if request.method == 'DELETE':
        try:
            # Check authorization
            if request.user.is_authenticated:
                try:
                    user_profile = UserProfile.objects.get(user=request.user)
                    if user_profile.role not in ['finance', 'admin', 'manager']:
                        return JsonResponse({
                            'success': False,
                            'message': 'Access Denied: Only Finance, Admin, or Manager can delete customers'
                        }, status=403)
                except UserProfile.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'message': 'User profile not found'
                    }, status=403)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Authentication required'
                }, status=401)
            
            customer = Customer.objects.get(id=customer_id)
            customer.delete()

            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='DELETE_CUSTOMER',
                table_name='customers',
                record_id=customer_id,
                old_values={'customer_code': customer.customer_code}
            )

            return JsonResponse({'success': True, 'message': 'Customer deleted successfully'})
        except Customer.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Customer not found'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)


@csrf_exempt
def api_dashboard_data(request):
    """API endpoint for dashboard data (for AJAX calls)"""
    if request.method == 'GET':
        try:
            # Get date ranges
            today = datetime.now().date()
            week_ago = today - timedelta(days=7)
            month_ago = today - timedelta(days=30)
            
            # Basic counts
            total_customers = Customer.objects.count()
            total_products = Product.objects.count()
            total_orders = Order.objects.count()
            
            # Today's sales
            today_sales = Order.objects.filter(
                order_date__date=today,
                status='delivered'
            ).aggregate(total=Sum('grand_total'))['total'] or 0
            
            # Month sales
            month_sales = Order.objects.filter(
                order_date__date__gte=month_ago,
                status='delivered'
            ).aggregate(total=Sum('grand_total'))['total'] or 0
            
            # Pending orders
            pending_orders = Order.objects.filter(status='pending').count()
            
            # Low stock
            low_stock = Product.objects.filter(stock_quantity__lt=F('min_stock_level')).count()
            
            # Recent orders
            recent_orders = Order.objects.select_related('customer__user').order_by('-order_date')[:5].values(
                'id', 'order_number', 'customer__user__username',
                'status', 'grand_total', 'order_date'
            )
            
            # Top products
            top_products = OrderItem.objects.select_related('product').values(
                'product__name'
            ).annotate(
                total_sold=Sum('quantity')
            ).order_by('-total_sold')[:5]
            
            data = {
                'stats': {
                    'total_customers': total_customers,
                    'total_products': total_products,
                    'total_orders': total_orders,
                    'today_sales': float(today_sales),
                    'month_sales': float(month_sales),
                    'pending_orders': pending_orders,
                    'low_stock': low_stock,
                },
                'recent_orders': list(recent_orders),
                'top_products': list(top_products)
            }
            
            return JsonResponse(data)
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid method'}, status=405)


@csrf_exempt
def api_add_customer(request):
    """Legacy: Add new customer"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Create user
            user = User.objects.create_user(
                username=data['username'],
                email=data.get('email', ''),
                first_name=data.get('first_name', ''),
                last_name=data.get('last_name', ''),
                password=data.get('password', 'password123')
            )
            
            # Create user profile
            profile = UserProfile.objects.create(
                user=user,
                unique_id=data.get('unique_id', f"CUST{user.id:04d}"),
                role='customer',
                phone=data.get('phone', ''),
                department='Customer'
            )
            
            # Create customer
            customer = Customer.objects.create(
                user=user,
                customer_code=data.get('customer_code', f"CUST{user.id:04d}"),
                billing_address=data.get('billing_address', ''),
                shipping_address=data.get('shipping_address', ''),
                credit_limit=data.get('credit_limit', 10000),
                tax_number=data.get('tax_number', '')
            )
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='CREATE_CUSTOMER',
                table_name='customers',
                record_id=customer.id,
                new_values={'customer_code': customer.customer_code}
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Customer added successfully',
                'customer_id': customer.id
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)


@csrf_exempt
def api_add_product(request):
    """Add new product"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Get or create category
            category = None
            if data.get('category'):
                category, _ = ProductCategory.objects.get_or_create(
                    name=data['category']
                )
            
            product = Product.objects.create(
                name=data['name'],
                sku=data['sku'],
                description=data.get('description', ''),
                category=category,
                price=data['price'],
                cost=data.get('cost', data['price'] * 0.7),
                stock_quantity=data.get('stock_quantity', 0),
                min_stock_level=data.get('min_stock_level', 10),
                is_active=data.get('is_active', True),
                created_by=request.user if request.user.is_authenticated else None
            )
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='CREATE_PRODUCT',
                table_name='products',
                record_id=product.id,
                new_values={'sku': product.sku, 'name': product.name}
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Product added successfully',
                'product_id': product.id
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)


@csrf_exempt
def api_create_order(request):
    """Create new order"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Get customer
            customer = Customer.objects.get(id=data['customer'])
            
            # Use provided values or generate
            order_number = data.get('order_number', f"ORD{Order.objects.count() + 1:05d}")
            total_amount = float(data.get('total_amount', 0))
            tax_amount = float(data.get('tax_amount', 0))
            discount_amount = float(data.get('discount_amount', 0))
            grand_total = float(data.get('grand_total', total_amount + tax_amount - discount_amount))
            
            # Create order
            order = Order.objects.create(
                order_number=order_number,
                customer=customer,
                total_amount=total_amount,
                tax_amount=tax_amount,
                discount_amount=discount_amount,
                grand_total=grand_total,
                status=data.get('status', 'pending'),
                payment_status=data.get('payment_status', 'pending'),
                shipping_address=data.get('shipping_address', customer.shipping_address),
                notes=data.get('notes', ''),
                created_by=request.user if request.user.is_authenticated else None
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Order created successfully',
                'order_number': order.order_number,
                'order_id': order.id
            })
        except Customer.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Customer not found'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({'error': 'Invalid method'}, status=405)


@csrf_exempt
def api_get_customers(request):
    """Legacy: Get all customers"""
    try:
        customers = Customer.objects.select_related('user').all().values(
            'id', 'customer_code', 'user__username', 
            'user__email', 'user__first_name', 'user__last_name',
            'credit_limit', 'balance', 'created_at'
        )
        customer_list = []
        for c in customers:
            customer_list.append({
                'id': c['id'],
                'customer_code': c['customer_code'],
                'user_email': c['user__email'],
                'user_first_name': c['user__first_name'],
                'user_last_name': c['user__last_name'],
                'phone': '',
                'credit_limit': float(c['credit_limit']),
                'balance': float(c['balance'])
            })
        return JsonResponse({'results': customer_list})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def api_get_products(request):
    """Legacy: Get all products"""
    try:
        products = Product.objects.select_related('category').all().values(
            'id', 'name', 'sku', 'category__name',
            'price', 'cost', 'stock_quantity', 'min_stock_level', 'is_active', 'created_at'
        )
        product_list = []
        for p in products:
            product_list.append({
                'id': p['id'],
                'name': p['name'],
                'sku': p['sku'],
                'category_name': p['category__name'] or 'Uncategorized',
                'price': float(p['price']),
                'cost': float(p['cost']) if p['cost'] else float(p['price']) * 0.7,
                'stock_quantity': p['stock_quantity'],
                'min_stock_level': p['min_stock_level'],
                'is_active': p['is_active']
            })
        return JsonResponse({'results': product_list})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def api_get_orders(request):
    """Legacy: Get all orders"""
    try:
        orders = Order.objects.select_related('customer__user').all().values(
            'id', 'order_number', 'customer__user__username',
            'customer__user__first_name', 'customer__user__last_name',
            'status', 'payment_status', 'grand_total', 'order_date'
        ).order_by('-order_date')
        order_list = []
        for o in orders:
            order_list.append({
                'id': o['id'],
                'order_number': o['order_number'],
                'customer_name': f"{o['customer__user__first_name']} {o['customer__user__last_name']}",
                'status': o['status'],
                'payment_status': o['payment_status'],
                'grand_total': float(o['grand_total']),
                'order_date': o['order_date']
            })
        return JsonResponse({'results': order_list})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)



# =============== TEST ENDPOINTS ===============
@api_view(['GET'])
@permission_classes([AllowAny])
def test_api(request):
    """Test endpoint to verify API is working"""
    return Response({
        'success': True,
        'message': 'API is working',
        'timestamp': datetime.now().isoformat(),
        'method': request.method,
        'params': dict(request.GET)
    })


@api_view(['GET'])
@permission_classes([AllowAny])
def test_customers(request):
    """Test customers endpoint"""
    try:
        count = Customer.objects.count()
        return Response({
            'success': True,
            'total_customers': count,
            'message': f'Found {count} customers in database'
        })
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


# =============== TEMPLATE VIEWS ===============
@login_required(login_url='/accounts/login/')
def dashboard_view(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


@login_required(login_url='/accounts/login/')
def crm_dashboard_view(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


@staff_required
def customers_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


@staff_required
def products_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


@staff_required
def orders_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


@staff_required
def inventory_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


@staff_required
def leads_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


@finance_required
def payments_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


@finance_required
def invoices_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


@finance_required
def reports_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


@login_required(login_url='/accounts/login/')
def settings_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


def help_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


def test_api_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


@finance_required
def activity_logs_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


def companies_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


@finance_required
def expenses_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


@admin_or_manager_required
def permissions_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


def product_categories_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


@staff_required
def product_tracking_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


def html_redirect(request, path):
    """Redirect .html requests to proper URL without .html"""
    # Remove .html extension and redirect
    clean_path = path.replace('.html', '')
    return redirect(f'/dashboard/{clean_path}/')


class CategoriesAPIView(APIView):
    """List and create product categories"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        """Get all product categories"""
        try:
            categories = ProductCategory.objects.all()
            
            results = []
            for category in categories:
                results.append({
                    'id': category.id,
                    'name': category.name,
                    'description': category.description or '',
                    'created_at': category.created_at.isoformat() if category.created_at else ''
                })
            
            return Response({
                'success': True,
                'results': results,
                'count': len(results)
            })
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def post(self, request):
        """Create new product category"""
        try:
            data = request.data
            
            # Validate required fields
            if not data.get('name'):
                return Response({
                    'success': False,
                    'error': 'Category name is required'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check if category already exists
            if ProductCategory.objects.filter(name__iexact=data['name']).exists():
                return Response({
                    'success': False,
                    'error': 'Category with this name already exists'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Create category
            category = ProductCategory.objects.create(
                name=data['name'].strip(),
                description=data.get('description', '').strip()
            )
            
            return Response({
                'success': True,
                'message': 'Category created successfully',
                'category_id': category.id,
                'category_name': category.name
            })
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# =============== USER MANAGEMENT VIEWS ===============
@login_required(login_url='/accounts/login/')
def get_staff_finance_users(request):
    """Get list of staff and finance users"""
    # Check if user is authenticated and is a manager/admin
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    try:
        profile = UserProfile.objects.get(user=request.user)
        if profile.role not in ['manager', 'admin']:
            return JsonResponse({'error': 'Permission denied. Only managers and admins can view users.'}, status=403)
    except UserProfile.DoesNotExist:
        return JsonResponse({'error': 'User profile not found'}, status=403)
    
    try:
        staff_users = UserProfile.objects.filter(role='staff').select_related('user')
        finance_users = UserProfile.objects.filter(role='finance').select_related('user')
        
        staff_list = [{
            'id': profile.user.id,
            'username': profile.user.username,
            'email': profile.user.email,
            'name': f"{profile.user.first_name} {profile.user.last_name}".strip(),
            'unique_id': profile.unique_id,
            'department': profile.department or 'N/A',
            'phone': profile.phone or 'N/A',
            'created_at': profile.user.date_joined.strftime('%Y-%m-%d %H:%M')
        } for profile in staff_users]
        
        finance_list = [{
            'id': profile.user.id,
            'username': profile.user.username,
            'email': profile.user.email,
            'name': f"{profile.user.first_name} {profile.user.last_name}".strip(),
            'unique_id': profile.unique_id,
            'department': profile.department or 'N/A',
            'phone': profile.phone or 'N/A',
            'created_at': profile.user.date_joined.strftime('%Y-%m-%d %H:%M')
        } for profile in finance_users]
        
        return JsonResponse({
            'success': True,
            'staff_users': staff_list,
            'finance_users': finance_list
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@csrf_exempt
def health_check(request):
    """Health check endpoint"""
    return JsonResponse({'status': 'ok', 'message': 'API is working'})


@csrf_exempt
def create_staff_user(request):
    """Create a new staff user"""
    print(f"\n{'='*70}")
    print(f"[DEBUG] create_staff_user called - Method: {request.method}, Authenticated: {request.user.is_authenticated}")
    print(f"[DEBUG] User: {request.user}, Session Key: {request.session.session_key}")
    print(f"[DEBUG] REQUEST.POST data: {dict(request.POST)}")
    print(f"[DEBUG] REQUEST.FILES data: {dict(request.FILES)}")
    print(f"[DEBUG] Content-Type: {request.META.get('CONTENT_TYPE')}")
    print(f"{'='*70}\n")
    
    # Check if user is authenticated and is a manager/admin
    if not request.user.is_authenticated:
        print(f"[ERROR] User not authenticated")
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    try:
        profile = UserProfile.objects.get(user=request.user)
        print(f"[DEBUG] User profile found - Role: {profile.role}")
        if profile.role not in ['manager', 'admin']:
            print(f"[ERROR] Permission denied for role: {profile.role}")
            return JsonResponse({'error': 'Permission denied. Only managers and admins can create users.'}, status=403)
    except UserProfile.DoesNotExist:
        print(f"[ERROR] UserProfile does not exist for user {request.user}")
        return JsonResponse({'error': 'User profile not found'}, status=403)
    
    if request.method != 'POST':
        print(f"[ERROR] Invalid method: {request.method}")
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = request.POST
        print(f"[DEBUG] Staff Form Data: {dict(data)}")
        
        # Validate required fields
        required = ['username', 'password', 'email', 'first_name']
        missing = [field for field in required if not data.get(field)]
        if missing:
            print(f"[DEBUG] Missing fields: {missing}")
            return JsonResponse({
                'error': f'Missing required fields: {", ".join(missing)}'
            }, status=400)
        
        # Check if username exists
        if User.objects.filter(username=data['username']).exists():
            return JsonResponse({
                'error': 'Username already exists'
            }, status=400)
        
        if User.objects.filter(email=data['email']).exists():
            return JsonResponse({
                'error': 'Email already exists'
            }, status=400)
        
        # Create user
        user = User.objects.create_user(
            username=data['username'],
            email=data['email'],
            password=data['password'],
            first_name=data.get('first_name', ''),
            last_name=data.get('last_name', '')
        )
        print(f"[DEBUG] Created user: {user.username} (ID: {user.id})")
        
        # Create profile
        unique_id = f"STAFF{user.id:04d}"
        counter = 1
        original_id = unique_id
        while UserProfile.objects.filter(unique_id=unique_id).exists():
            unique_id = f"{original_id}_{counter}"
            counter += 1
        
        profile, created = UserProfile.objects.get_or_create(
            user=user,
            defaults={
                'unique_id': unique_id,
                'role': 'staff',
                'department': data.get('department', ''),
                'phone': data.get('phone', '')
            }
        )
        
        # If profile already existed, update it
        if not created:
            profile.unique_id = unique_id
            profile.role = 'staff'
            profile.department = data.get('department', '')
            profile.phone = data.get('phone', '')
            profile.save()
        
        print(f"[DEBUG] Created profile with ID: {profile.unique_id}")
        
        return JsonResponse({
            'success': True,
            'message': 'Staff user created successfully',
            'user_id': user.id,
            'username': user.username,
            'unique_id': profile.unique_id
        })
    except Exception as e:
        print(f"[ERROR] Staff user creation failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@csrf_exempt
def create_finance_user(request):
    """Create a new finance user"""
    # Check if user is authenticated and is a manager/admin
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    try:
        profile = UserProfile.objects.get(user=request.user)
        if profile.role not in ['manager', 'admin']:
            return JsonResponse({'error': 'Permission denied. Only managers and admins can create users.'}, status=403)
    except UserProfile.DoesNotExist:
        return JsonResponse({'error': 'User profile not found'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = request.POST
        print(f"[DEBUG] Finance Form Data: {dict(data)}")
        
        # Validate required fields
        required = ['username', 'password', 'email', 'first_name']
        missing = [field for field in required if not data.get(field)]
        if missing:
            print(f"[DEBUG] Missing fields: {missing}")
            return JsonResponse({
                'error': f'Missing required fields: {", ".join(missing)}'
            }, status=400)
        
        # Check if username exists
        if User.objects.filter(username=data['username']).exists():
            return JsonResponse({
                'error': 'Username already exists'
            }, status=400)
        
        if User.objects.filter(email=data['email']).exists():
            return JsonResponse({
                'error': 'Email already exists'
            }, status=400)
        
        # Create user
        user = User.objects.create_user(
            username=data['username'],
            email=data['email'],
            password=data['password'],
            first_name=data.get('first_name', ''),
            last_name=data.get('last_name', '')
        )
        print(f"[DEBUG] Created user: {user.username} (ID: {user.id})")
        
        # Create profile
        unique_id = f"FIN{user.id:04d}"
        counter = 1
        original_id = unique_id
        while UserProfile.objects.filter(unique_id=unique_id).exists():
            unique_id = f"{original_id}_{counter}"
            counter += 1
        
        profile, created = UserProfile.objects.get_or_create(
            user=user,
            defaults={
                'unique_id': unique_id,
                'role': 'finance',
                'department': data.get('department', ''),
                'phone': data.get('phone', '')
            }
        )
        
        # If profile already existed, update it
        if not created:
            profile.unique_id = unique_id
            profile.role = 'finance'
            profile.department = data.get('department', '')
            profile.phone = data.get('phone', '')
            profile.save()
        
        print(f"[DEBUG] Created profile with ID: {profile.unique_id}")
        
        return JsonResponse({
            'success': True,
            'message': 'Finance user created successfully',
            'user_id': user.id,
            'username': user.username,
            'unique_id': profile.unique_id
        })
    except Exception as e:
        print(f"[ERROR] Finance user creation failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required(login_url='/accounts/login/')
def delete_user(request, user_id):
    """Delete a staff or finance user"""
    # Check if user is authenticated and is a manager/admin
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    try:
        profile = UserProfile.objects.get(user=request.user)
        if profile.role not in ['manager', 'admin']:
            return JsonResponse({'error': 'Permission denied. Only managers and admins can delete users.'}, status=403)
    except UserProfile.DoesNotExist:
        return JsonResponse({'error': 'User profile not found'}, status=403)
    
    if request.method != 'DELETE' and request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        user = User.objects.get(id=user_id)
        profile = UserProfile.objects.get(user=user)
        
        # Only allow deletion of staff and finance users
        if profile.role not in ['staff', 'finance']:
            return JsonResponse({
                'error': 'Can only delete staff or finance users'
            }, status=400)
        
        username = user.username
        unique_id = profile.unique_id
        
        # Delete user (cascades to profile)
        user.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'User {username} ({unique_id}) deleted successfully'
        })
    except User.DoesNotExist:
        return JsonResponse({
            'error': 'User not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

# =============== CMS MANAGEMENT VIEWS ===============

@login_required(login_url='/accounts/login/')
def cms_dashboard(request):
    """CMS Management Dashboard in CRM"""
    context = {
        'total_content': CMSContent.objects.count(),
        'active_content': CMSContent.objects.filter(is_active=True).count(),
        'total_pages': CMSPage.objects.count(),
        'published_pages': CMSPage.objects.filter(is_published=True).count(),
        'recent_content': CMSContent.objects.all()[:5],
        'recent_pages': CMSPage.objects.all()[:5],
    }
    return render(request, 'cms_dashboard.html', context)

@login_required(login_url='/accounts/login/')
def cms_content_list(request):
    """List all CMS content items"""
    content_type = request.GET.get('type', '')
    search = request.GET.get('search', '')
    
    items = CMSContent.objects.all()
    
    if content_type:
        items = items.filter(content_type=content_type)
    
    if search:
        items = items.filter(Q(title__icontains=search) | Q(description__icontains=search))
    
    items = items.order_by('-created_at')
    
    content_types = [
        ('banner', 'Banner'),
        ('section', 'Section'),
        ('page', 'Page'),
        ('testimonial', 'Testimonial'),
        ('feature', 'Feature'),
        ('announcement', 'Announcement'),
    ]
    
    context = {
        'items': items,
        'content_types': content_types,
        'selected_type': content_type,
        'search_query': search,
        'total': items.count(),
    }
    return render(request, 'cms_content_list.html', context)

@login_required(login_url='/accounts/login/')
def cms_add_content(request):
    """Add new CMS content"""
    if request.method == 'POST':
        title = request.POST.get('title', '')
        slug = request.POST.get('slug', '')
        content_type = request.POST.get('content_type', 'section')
        description = request.POST.get('description', '')
        order = int(request.POST.get('order', 0))
        is_active = request.POST.get('is_active') == 'on'
        is_featured = request.POST.get('is_featured') == 'on'
        image = request.FILES.get('image', None)
        
        # Validate
        if not title or not content_type:
            return render(request, 'cms_add_content.html', {
                'error': 'Title and Content Type are required',
                'form_data': request.POST
            })
        
        # Auto-generate slug if empty
        if not slug:
            slug = title.lower().replace(' ', '-').replace('_', '-')
        
        # Check slug uniqueness
        if CMSContent.objects.filter(slug=slug).exists():
            return render(request, 'cms_add_content.html', {
                'error': 'Slug already exists. Please use a different slug.',
                'form_data': request.POST
            })
        
        # Create content
        content = CMSContent.objects.create(
            title=title,
            slug=slug,
            content_type=content_type,
            description=description,
            order=order,
            is_active=is_active,
            is_featured=is_featured,
            author=request.user,
            image=image,
            published_at=timezone.now() if is_active else None
        )
        
        # Log activity
        ActivityLog.objects.create(
            user=request.user,
            action='CREATE',
            table_name='CMS Content',
            record_id=content.id,
            new_values={'title': title, 'type': content_type}
        )
        
        return redirect('cms_view_content', content_id=content.id)
    
    context = {
        'content_types': [
            ('banner', 'Banner'),
            ('section', 'Section'),
            ('page', 'Page'),
            ('testimonial', 'Testimonial'),
            ('feature', 'Feature'),
            ('announcement', 'Announcement'),
        ]
    }
    return render(request, 'cms_add_content.html', context)

@login_required(login_url='/accounts/login/')
def cms_view_content(request, content_id):
    """View/Edit CMS content"""
    try:
        content = CMSContent.objects.get(id=content_id)
    except CMSContent.DoesNotExist:
        return render(request, '404.html', {'message': 'Content not found'}, status=404)
    
    if request.method == 'POST':
        # Update content
        old_values = {
            'title': content.title,
            'description': content.description[:100],
        }
        
        content.title = request.POST.get('title', content.title)
        content.slug = request.POST.get('slug', content.slug)
        content.description = request.POST.get('description', content.description)
        content.order = int(request.POST.get('order', content.order))
        content.is_active = request.POST.get('is_active') == 'on'
        content.is_featured = request.POST.get('is_featured') == 'on'
        
        if request.FILES.get('image'):
            content.image = request.FILES.get('image')
        
        content.save()
        
        # Log activity
        ActivityLog.objects.create(
            user=request.user,
            action='UPDATE',
            table_name='CMS Content',
            record_id=content.id,
            old_values=old_values,
            new_values={'title': content.title}
        )
        
        return redirect('cms_view_content', content_id=content.id)
    
    context = {
        'content': content,
        'content_types': [
            ('banner', 'Banner'),
            ('section', 'Section'),
            ('page', 'Page'),
            ('testimonial', 'Testimonial'),
            ('feature', 'Feature'),
            ('announcement', 'Announcement'),
        ]
    }
    return render(request, 'cms_view_content.html', context)

@login_required(login_url='/accounts/login/')
def cms_delete_content(request, content_id):
    """Delete CMS content"""
    try:
        content = CMSContent.objects.get(id=content_id)
        title = content.title
        content.delete()
        
        # Log activity
        ActivityLog.objects.create(
            user=request.user,
            action='DELETE',
            table_name='CMS Content',
            record_id=content_id,
            old_values={'title': title}
        )
        
        return redirect('cms_content_list')
    except CMSContent.DoesNotExist:
        return JsonResponse({'error': 'Content not found'}, status=404)

@login_required(login_url='/accounts/login/')
def cms_pages_list(request):
    """List all CMS pages"""
    search = request.GET.get('search', '')
    
    pages = CMSPage.objects.all()
    
    if search:
        pages = pages.filter(Q(title__icontains=search) | Q(slug__icontains=search))
    
    pages = pages.order_by('-created_at')
    
    context = {
        'pages': pages,
        'search_query': search,
        'total': pages.count(),
    }
    return render(request, 'cms_pages_list.html', context)

@login_required(login_url='/accounts/login/')
def cms_add_page(request):
    """Add new CMS page"""
    if request.method == 'POST':
        title = request.POST.get('title', '')
        slug = request.POST.get('slug', '')
        page_title = request.POST.get('page_title', '')
        page_description = request.POST.get('page_description', '')
        meta_keywords = request.POST.get('meta_keywords', '')
        is_published = request.POST.get('is_published') == 'on'
        hero_image = request.FILES.get('hero_image', None)
        
        if not title:
            return render(request, 'cms_add_page.html', {
                'error': 'Title is required',
            })
        
        if not slug:
            slug = title.lower().replace(' ', '-')
        
        if CMSPage.objects.filter(slug=slug).exists():
            return render(request, 'cms_add_page.html', {
                'error': 'Slug already exists',
            })
        
        page = CMSPage.objects.create(
            title=title,
            slug=slug,
            page_title=page_title,
            page_description=page_description,
            meta_keywords=meta_keywords,
            is_published=is_published,
            hero_image=hero_image,
            author=request.user,
            published_at=timezone.now() if is_published else None
        )
        
        ActivityLog.objects.create(
            user=request.user,
            action='CREATE',
            table_name='CMS Page',
            record_id=page.id,
            new_values={'title': title}
        )
        
        return redirect('cms_view_page', page_id=page.id)
    
    return render(request, 'cms_add_page.html')

@login_required(login_url='/accounts/login/')
def cms_view_page(request, page_id):
    """View/Edit CMS page"""
    try:
        page = CMSPage.objects.get(id=page_id)
    except CMSPage.DoesNotExist:
        return render(request, '404.html', {'message': 'Page not found'}, status=404)
    
    if request.method == 'POST':
        page.title = request.POST.get('title', page.title)
        page.slug = request.POST.get('slug', page.slug)
        page.page_title = request.POST.get('page_title', page.page_title)
        page.page_description = request.POST.get('page_description', page.page_description)
        page.meta_keywords = request.POST.get('meta_keywords', page.meta_keywords)
        page.is_published = request.POST.get('is_published') == 'on'
        
        if request.FILES.get('hero_image'):
            page.hero_image = request.FILES.get('hero_image')
        
        page.save()
        
        ActivityLog.objects.create(
            user=request.user,
            action='UPDATE',
            table_name='CMS Page',
            record_id=page.id,
            new_values={'title': page.title}
        )
        
        return redirect('cms_view_page', page_id=page.id)
    
    sections = page.sections.all()
    available_content = CMSContent.objects.filter(is_active=True).exclude(
        id__in=sections.values_list('content_id', flat=True)
    )
    
    context = {
        'page': page,
        'sections': sections,
        'available_content': available_content,
    }
    return render(request, 'cms_view_page.html', context)

@login_required(login_url='/accounts/login/')
def cms_add_section(request, page_id):
    """Add content section to page"""
    try:
        page = CMSPage.objects.get(id=page_id)
    except CMSPage.DoesNotExist:
        return JsonResponse({'error': 'Page not found'}, status=404)
    
    if request.method == 'POST':
        content_id = request.POST.get('content_id')
        order = int(request.POST.get('order', 0))
        
        try:
            content = CMSContent.objects.get(id=content_id)
        except CMSContent.DoesNotExist:
            return JsonResponse({'error': 'Content not found'}, status=404)
        
        section = CMSPageSection.objects.create(
            page=page,
            content=content,
            order=order
        )
        
        return redirect('cms_view_page', page_id=page.id)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@login_required(login_url='/accounts/login/')
def cms_delete_section(request, section_id):
    """Delete page section"""
    try:
        section = CMSPageSection.objects.get(id=section_id)
        page_id = section.page.id
        section.delete()
        
        return redirect('cms_view_page', page_id=page_id)
    except CMSPageSection.DoesNotExist:
        return JsonResponse({'error': 'Section not found'}, status=404)

@login_required(login_url='/accounts/login/')
def cms_delete_page(request, page_id):
    """Delete CMS page"""
    try:
        page = CMSPage.objects.get(id=page_id)
        title = page.title
        page.delete()
        
        ActivityLog.objects.create(
            user=request.user,
            action='DELETE',
            table_name='CMS Page',
            record_id=page_id,
            old_values={'title': title}
        )
        
        return redirect('cms_pages_list')
    except CMSPage.DoesNotExist:
        return JsonResponse({'error': 'Page not found'}, status=404)

# ============= WEBSITE HERO HEADERS API =============

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def hero_headers_list(request):
    """List all hero headers or create new"""
    if request.method == 'GET':
        headers = WebsiteHeroHeader.objects.all()
        serializer = WebsiteHeroHeaderSerializer(headers, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        serializer = WebsiteHeroHeaderSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=request.user)
            ActivityLog.objects.create(
                user=request.user,
                action='CREATE',
                table_name='Website Hero Headers',
                record_id=serializer.instance.id,
                new_values=serializer.data
            )
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def hero_header_detail(request, header_id):
    """Get, update, or delete specific hero header"""
    try:
        header = WebsiteHeroHeader.objects.get(id=header_id)
    except WebsiteHeroHeader.DoesNotExist:
        return Response({'error': 'Hero header not found'}, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        serializer = WebsiteHeroHeaderSerializer(header)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        old_values = WebsiteHeroHeaderSerializer(header).data
        serializer = WebsiteHeroHeaderSerializer(header, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            ActivityLog.objects.create(
                user=request.user,
                action='UPDATE',
                table_name='Website Hero Headers',
                record_id=header.id,
                old_values=old_values,
                new_values=serializer.data
            )
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        old_values = WebsiteHeroHeaderSerializer(header).data
        header.delete()
        ActivityLog.objects.create(
            user=request.user,
            action='DELETE',
            table_name='Website Hero Headers',
            record_id=header_id,
            old_values=old_values
        )
        return Response({'message': 'Hero header deleted'}, status=status.HTTP_204_NO_CONTENT)

# ============= WEBSITE CONDUCTS API =============

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def conducts_list(request):
    """List all website conducts or create new"""
    if request.method == 'GET':
        conducts = WebsiteConduct.objects.filter(is_active=True).order_by('order')
        serializer = WebsiteConductSerializer(conducts, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        serializer = WebsiteConductSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=request.user)
            ActivityLog.objects.create(
                user=request.user,
                action='CREATE',
                table_name='Website Conducts',
                record_id=serializer.instance.id,
                new_values=serializer.data
            )
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def conduct_detail(request, conduct_id):
    """Get, update, or delete specific conduct"""
    try:
        conduct = WebsiteConduct.objects.get(id=conduct_id)
    except WebsiteConduct.DoesNotExist:
        return Response({'error': 'Conduct not found'}, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        serializer = WebsiteConductSerializer(conduct)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        old_values = WebsiteConductSerializer(conduct).data
        serializer = WebsiteConductSerializer(conduct, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            ActivityLog.objects.create(
                user=request.user,
                action='UPDATE',
                table_name='Website Conducts',
                record_id=conduct.id,
                old_values=old_values,
                new_values=serializer.data
            )
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        old_values = WebsiteConductSerializer(conduct).data
        conduct.delete()
        ActivityLog.objects.create(
            user=request.user,
            action='DELETE',
            table_name='Website Conducts',
            record_id=conduct_id,
            old_values=old_values
        )
        return Response({'message': 'Conduct deleted'}, status=status.HTTP_204_NO_CONTENT)

# ============= WEBSITE PRODUCT DISPLAY API =============

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def product_displays_list(request):
    """List all website product displays or create new"""
    if request.method == 'GET':
        displays = WebsiteProductDisplay.objects.filter(is_displayed=True).order_by('-is_featured', 'display_order')
        serializer = WebsiteProductDisplaySerializer(displays, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        serializer = WebsiteProductDisplaySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=request.user)
            ActivityLog.objects.create(
                user=request.user,
                action='CREATE',
                table_name='Website Product Displays',
                record_id=serializer.instance.id,
                new_values=serializer.data
            )
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def product_display_detail(request, display_id):
    """Get, update, or delete specific product display"""
    try:
        display = WebsiteProductDisplay.objects.get(id=display_id)
    except WebsiteProductDisplay.DoesNotExist:
        return Response({'error': 'Product display not found'}, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        serializer = WebsiteProductDisplaySerializer(display)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        old_values = WebsiteProductDisplaySerializer(display).data
        serializer = WebsiteProductDisplaySerializer(display, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            ActivityLog.objects.create(
                user=request.user,
                action='UPDATE',
                table_name='Website Product Displays',
                record_id=display.id,
                old_values=old_values,
                new_values=serializer.data
            )
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        old_values = WebsiteProductDisplaySerializer(display).data
        display.delete()
        ActivityLog.objects.create(
            user=request.user,
            action='DELETE',
            table_name='Website Product Displays',
            record_id=display_id,
            old_values=old_values
        )
        return Response({'message': 'Product display deleted'}, status=status.HTTP_204_NO_CONTENT)

# ============= WEBSITE ENQUIRY API =============

@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
def website_enquiries_create(request):
    """Create website enquiry or list all (for authenticated users)"""
    if request.method == 'POST':
        serializer = WebsiteEnquirySerializer(data=request.data)
        if serializer.is_valid():
            # Generate enquiry number
            count = WebsiteEnquiry.objects.count() + 1
            enquiry_number = f"ENQ{timezone.now().strftime('%Y%m%d')}{count:04d}"
            
            enquiry = serializer.save(
                enquiry_number=enquiry_number,
                ip_address=get_client_ip(request)
            )
            
            # Log for admins
            ActivityLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='WEBSITE_ENQUIRY',
                table_name='Website Enquiries',
                record_id=enquiry.id,
                new_values=WebsiteEnquirySerializer(enquiry).data,
                ip_address=get_client_ip(request)
            )
            
            return Response({
                'message': 'Enquiry submitted successfully',
                'enquiry_number': enquiry_number,
                'data': serializer.data
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'GET':
        if not request.user.is_authenticated:
            return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)
        
        enquiries = WebsiteEnquiry.objects.all()
        serializer = WebsiteEnquirySerializer(enquiries, many=True)
        return Response(serializer.data)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def website_enquiry_detail(request, enquiry_id):
    """Get, update, or delete website enquiry"""
    try:
        enquiry = WebsiteEnquiry.objects.get(id=enquiry_id)
    except WebsiteEnquiry.DoesNotExist:
        return Response({'error': 'Enquiry not found'}, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        serializer = WebsiteEnquirySerializer(enquiry)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        old_values = WebsiteEnquirySerializer(enquiry).data
        serializer = WebsiteEnquirySerializer(enquiry, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            ActivityLog.objects.create(
                user=request.user,
                action='UPDATE',
                table_name='Website Enquiries',
                record_id=enquiry.id,
                old_values=old_values,
                new_values=serializer.data
            )
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        old_values = WebsiteEnquirySerializer(enquiry).data
        enquiry.delete()
        ActivityLog.objects.create(
            user=request.user,
            action='DELETE',
            table_name='Website Enquiries',
            record_id=enquiry_id,
            old_values=old_values
        )
        return Response({'message': 'Enquiry deleted'}, status=status.HTTP_204_NO_CONTENT)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def convert_enquiry_to_lead(request, enquiry_id):
    """Convert website enquiry to CRM lead"""
    try:
        enquiry = WebsiteEnquiry.objects.get(id=enquiry_id)
    except WebsiteEnquiry.DoesNotExist:
        return Response({'error': 'Enquiry not found'}, status=status.HTTP_404_NOT_FOUND)
    
    try:
        # Create lead from enquiry
        lead_count = Lead.objects.count() + 1
        lead_number = f"LEAD{timezone.now().strftime('%Y%m%d')}{lead_count:04d}"
        
        lead = Lead.objects.create(
            lead_number=lead_number,
            company_name=enquiry.company_name,
            contact_person=enquiry.contact_person,
            email=enquiry.email,
            phone=enquiry.phone,
            source='website',
            status='new',
            notes=f"Website Enquiry: {enquiry.subject}\n{enquiry.message}",
            created_by=request.user
        )
        
        # Update enquiry
        enquiry.lead = lead
        enquiry.status = 'converted'
        enquiry.converted_at = timezone.now()
        enquiry.converted_by = request.user
        enquiry.save()
        
        ActivityLog.objects.create(
            user=request.user,
            action='CONVERT_ENQUIRY',
            table_name='Website Enquiries',
            record_id=enquiry_id,
            new_values={'converted_to_lead': lead_number}
        )
        
        return Response({
            'message': 'Enquiry converted to lead',
            'lead_number': lead_number,
            'lead': {
                'id': lead.id,
                'lead_number': lead.lead_number,
                'company_name': lead.company_name
            }
        }, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ============= CMS MANAGEMENT PAGES (CRM Dashboard) =============

@login_required(login_url='/accounts/login/')
def cms_hero_headers_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


@login_required(login_url='/accounts/login/')
def cms_conducts_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


@login_required(login_url='/accounts/login/')
def cms_products_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


@login_required(login_url='/accounts/login/')
def cms_enquiries_page(request):
    """Redirect to React frontend - API only backend"""
    from django.shortcuts import redirect
    return redirect('http://localhost:3003/')


# ============= HOMEPAGE VIEW WITH ADMIN-MANAGED CONTENT =============

def homepage_dynamic(request):
    """Dynamic homepage using admin-managed content from models"""
    try:
        # Get hero section
        hero_section = HomepageHeroSection.objects.filter(is_active=True).first()
        
        # Get features
        features = HomepageFeature.objects.filter(is_active=True).order_by('order')
        
        # Get products
        products = Product.objects.filter(is_active=True)[:12]
        
        # Get sections
        why_us_section = HomepageSection.objects.filter(section_type='why_us', is_active=True).first()
        details_section = HomepageSection.objects.filter(section_type='details', is_active=True).first()
        stories_section = HomepageSection.objects.filter(section_type='stories', is_active=True).first()
        instagram_section = HomepageSection.objects.filter(section_type='instagram', is_active=True).first()
        testimonials_section = HomepageSection.objects.filter(section_type='testimonials', is_active=True).first()
        
        # Get navigation items
        navigation_items = HomepageNavigation.objects.filter(is_active=True).order_by('order')
        
        # Get footer sections
        footer_sections = HomepageFooterSection.objects.filter(is_active=True).order_by('order')
        
        # Get social links
        social_links = HomepageSocialLink.objects.filter(is_active=True).order_by('order')
        
        # Get SEO data
        seo = HomepageSEO.objects.first()
        
        context = {
            'hero_section': hero_section,
            'features': features,
            'products': products,
            'why_us_section': why_us_section,
            'details_section': details_section,
            'stories_section': stories_section,
            'instagram_section': instagram_section,
            'testimonials_section': testimonials_section,
            'navigation_items': navigation_items,
            'footer_sections': footer_sections,
            'social_links': social_links,
            'seo': seo,
        }
        
        return render(request, 'website/index_dynamic.html', context)
    
    except Exception as e:
        # Fallback to static version if error occurs
        print(f"Error rendering dynamic homepage: {str(e)}")
        products = Product.objects.filter(is_active=True)[:12]
        context = {'products': products}
        return render(request, 'website/index.html', context)


# ============= HELPER FUNCTIONS =============

def get_client_ip(request):
    """Get client IP address from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


# ============= HOMEPAGE CMS API VIEWS (FOR REACT FRONTEND) =============

class HomepageHeroAPIView(APIView):
    """Get homepage hero section data"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        try:
            hero = HomepageHeroSection.objects.filter(is_active=True).first()
            if hero:
                data = {
                    'id': hero.id,
                    'heading': hero.heading,
                    'subheading': hero.subheading,
                    'background_image': hero.background_image.url if hero.background_image else None,
                    'background_color': hero.background_color,
                    'cta_button_text': hero.cta_button_text,
                    'cta_button_url': hero.cta_button_url,
                    'text_color': hero.text_color,
                }
                return Response(data, status=status.HTTP_200_OK)
            return Response(None, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class HomepageFeaturesAPIView(APIView):
    """Get homepage features"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        try:
            features = HomepageFeature.objects.filter(is_active=True).order_by('order')
            data = []
            for feature in features:
                data.append({
                    'id': feature.id,
                    'title': feature.title,
                    'description': feature.description,
                    'icon_type': feature.icon_type,
                    'icon_image': feature.icon_image.url if feature.icon_image else None,
                    'background_color': feature.background_color,
                    'text_color': feature.text_color,
                    'accent_color': feature.accent_color,
                })
            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request):
        try:
            data = request.data
            feature = HomepageFeature.objects.create(
                title=data.get('title', ''),
                description=data.get('description', ''),
                icon_type=data.get('icon_type', ''),
                background_color=data.get('background_color', '#ffffff'),
                text_color=data.get('text_color', '#000000'),
                accent_color=data.get('accent_color', '#0000ff'),
                is_active=data.get('is_active', True),
            )
            return Response({'id': feature.id, 'message': 'Feature created'}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    """Get why us section with items"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        try:
            section = HomepageSection.objects.filter(section_type='why_us', is_active=True).first()
            if section:
                items = section.why_us_items.filter(is_active=True).order_by('order')
                data = {
                    'id': section.id,
                    'heading': section.heading,
                    'description': section.description,
                    'background_image': section.background_image.url if section.background_image else None,
                    'background_color': section.background_color,
                    'items': [{'id': item.id, 'text': item.text} for item in items],
                }
                return Response(data, status=status.HTTP_200_OK)
            return Response(None, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class HomepageDetailsAPIView(APIView):
    """Get details section with cards"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        try:
            section = HomepageSection.objects.filter(section_type='details', is_active=True).first()
            if section:
                cards = section.detail_cards.filter(is_active=True).order_by('order')
                data = {
                    'id': section.id,
                    'heading': section.heading,
                    'description': section.description,
                    'background_color': section.background_color,
                    'cards': [
                        {
                            'id': card.id,
                            'title': card.title,
                            'description': card.description,
                            'icon_type': card.icon_type,
                            'icon_image': card.icon_image.url if card.icon_image else None,
                        }
                        for card in cards
                    ],
                }
                return Response(data, status=status.HTTP_200_OK)
            return Response(None, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class HomepageStoriesAPIView(APIView):
    """Get stories section"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        try:
            section = HomepageSection.objects.filter(section_type='stories', is_active=True).first()
            stories = HomepageStory.objects.filter(is_active=True).order_by('-story_date')[:6]
            
            data = [
                {
                    'id': story.id,
                    'title': story.title,
                    'excerpt': story.excerpt,
                    'featured_image': story.featured_image.url if story.featured_image else None,
                    'story_date': story.story_date.isoformat(),
                    'read_more_url': story.read_more_url,
                    'icon_type': story.icon_type,
                }
                for story in stories
            ]
            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class HomepageInstagramAPIView(APIView):
    """Get Instagram section"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        try:
            instagram = HomepageInstagramSection.objects.filter(is_active=True).first()
            if instagram:
                data = {
                    'id': instagram.id,
                    'instagram_handle': instagram.instagram_handle,
                    'heading': instagram.heading,
                    'description': instagram.description,
                }
                return Response(data, status=status.HTTP_200_OK)
            return Response(None, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class HomepageTestimonialsAPIView(APIView):
    """Get testimonials section"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        try:
            testimonials = HomepageTestimonial.objects.filter(is_active=True).order_by('order')
            
            data = [
                {
                    'id': testimonial.id,
                    'testimonial_text': testimonial.testimonial_text,
                    'author_name': testimonial.author_name,
                    'author_title': testimonial.author_title,
                    'rating': testimonial.rating,
                }
                for testimonial in testimonials
            ]
            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request):
        try:
            data = request.data
            section = HomepageSection.objects.filter(section_type='testimonials').first()
            if not section:
                section = HomepageSection.objects.create(section_type='testimonials', is_active=True)
            testimonial = HomepageTestimonial.objects.create(
                section=section,
                testimonial_text=data.get('testimonial_text', ''),
                author_name=data.get('author_name', ''),
                author_title=data.get('author_title', 'Verified Purchase'),
                rating=int(data.get('rating', 5)),
                is_active=data.get('is_active', True),
            )
            return Response({'id': testimonial.id, 'message': 'Testimonial created'}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class HomepageNavigationAPIView(APIView):
    """Get navigation items"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        try:
            navigation = HomepageNavigation.objects.filter(is_active=True).order_by('order')
            
            data = [
                {
                    'id': nav.id,
                    'label': nav.label,
                    'url': nav.url,
                    'parent': nav.parent.id if nav.parent else None,
                    'order': nav.order,
                }
                for nav in navigation
            ]
            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request):
        try:
            data = request.data
            nav = HomepageNavigation.objects.create(
                label=data.get('label', ''),
                url=data.get('url', ''),
                order=int(data.get('order', 0)),
                is_active=data.get('is_active', True),
            )
            return Response({'id': nav.id, 'message': 'Navigation item created'}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class HomepageFooterAPIView(APIView):
    """Get footer sections"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        try:
            sections = HomepageFooterSection.objects.filter(is_active=True).order_by('order')
            
            data = []
            for section in sections:
                links = section.links.filter(is_active=True).order_by('order')
                section_data = {
                    'id': section.id,
                    'column_title': section.column_title,
                    'column_type': section.column_type,
                    'content': section.content,
                    'order': section.order,
                    'links': [
                        {
                            'id': link.id,
                            'link_text': link.link_text,
                            'link_url': link.link_url,
                        }
                        for link in links
                    ],
                }
                data.append(section_data)
            
            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request):
        try:
            data = request.data
            section = HomepageFooterSection.objects.create(
                column_title=data.get('column_title', ''),
                column_type=data.get('column_type', 'text'),
                content=data.get('content', ''),
                order=int(data.get('order', 0)),
                is_active=data.get('is_active', True),
            )
            return Response({'id': section.id, 'message': 'Footer section created'}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class HomepageSocialAPIView(APIView):
    """Get social media links"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        try:
            social = HomepageSocialLink.objects.filter(is_active=True).order_by('order')
            
            data = [
                {
                    'id': link.id,
                    'platform': link.platform,
                    'url': link.url,
                    'icon_class': link.icon_class,
                    'order': link.order,
                }
                for link in social
            ]
            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request):
        try:
            data = request.data
            link = HomepageSocialLink.objects.create(
                platform=data.get('platform', ''),
                url=data.get('url', ''),
                icon_class=data.get('icon_class', ''),
                order=int(data.get('order', 0)),
                is_active=data.get('is_active', True),
            )
            return Response({'id': link.id, 'message': 'Social link created'}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class HomepageSEOAPIView(APIView):
    """Get SEO data"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        try:
            seo = HomepageSEO.objects.first()
            if seo:
                data = {
                    'id': seo.id,
                    'page_title': seo.page_title,
                    'meta_description': seo.meta_description,
                    'meta_keywords': seo.meta_keywords,
                    'og_title': seo.og_title,
                    'og_description': seo.og_description,
                    'og_image': seo.og_image.url if seo.og_image else None,
                    'canonical_url': seo.canonical_url,
                }
                return Response(data, status=status.HTTP_200_OK)
            return Response(None, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request):
        try:
            data = request.data
            seo, created = HomepageSEO.objects.get_or_create(id=1)
            seo.page_title = data.get('page_title', '')
            seo.meta_description = data.get('meta_description', '')
            seo.meta_keywords = data.get('meta_keywords', '')
            seo.og_title = data.get('og_title', '')
            seo.og_description = data.get('og_description', '')
            seo.canonical_url = data.get('canonical_url', '')
            seo.save()
            return Response({'id': seo.id, 'message': 'SEO data updated'}, status=status.HTTP_200_OK if not created else status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# ============= SITE INFO API =============

class SiteInfoView(generics.RetrieveUpdateAPIView):
    queryset = SiteInfo.objects.all()
    serializer_class = SiteInfoSerializer
    permission_classes = [AllowAny]

    @csrf_exempt
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_object(self):
        return SiteInfo.objects.first()


# ============= USER PROFILE API =============

class UserProfileAPIView(APIView):
    """Get current user profile"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            user = request.user
            profile = UserProfile.objects.get(user=user)
            data = {
                'id': profile.id,
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'role': profile.role,
                'phone': profile.phone,
                'department': profile.department,
            }
            return Response(data, status=status.HTTP_200_OK)
        except UserProfile.DoesNotExist:
            return Response({'error': 'User profile not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ============= WEBSITE CONTENT APIs =============

class WebsiteStoryListCreateView(generics.ListCreateAPIView):
    """Get all stories or create a new story"""
    queryset = WebsiteStory.objects.filter(is_active=True).order_by('order')
    serializer_class = WebsiteStorySerializer
    permission_classes = [AllowAny]


class WebsiteStoryDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Get, update, or delete a story"""
    queryset = WebsiteStory.objects.all()
    serializer_class = WebsiteStorySerializer
    permission_classes = [AllowAny]


class WebsiteTestimonialListCreateView(generics.ListCreateAPIView):
    """Get all testimonials or create a new one"""
    queryset = WebsiteTestimonial.objects.filter(is_active=True).order_by('-created_at')
    serializer_class = WebsiteTestimonialSerializer
    permission_classes = [AllowAny]


class WebsiteTestimonialDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Get, update, or delete a testimonial"""
    queryset = WebsiteTestimonial.objects.all()
    serializer_class = WebsiteTestimonialSerializer
    permission_classes = [AllowAny]


class WebsiteGalleryListCreateView(generics.ListCreateAPIView):
    """Get all gallery items or create a new one"""
    queryset = WebsiteGallery.objects.filter(is_active=True).order_by('order')
    serializer_class = WebsiteGallerySerializer
    permission_classes = [AllowAny]


class WebsiteGalleryDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Get, update, or delete a gallery item"""
    queryset = WebsiteGallery.objects.all()
    serializer_class = WebsiteGallerySerializer
    permission_classes = [AllowAny]


class WebsiteFAQListCreateView(generics.ListCreateAPIView):
    """Get all FAQs or create a new one"""
    queryset = WebsiteFAQ.objects.filter(is_active=True).order_by('order')
    serializer_class = WebsiteFAQSerializer
    permission_classes = [AllowAny]


class WebsiteFAQDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Get, update, or delete an FAQ"""
    queryset = WebsiteFAQ.objects.all()
    serializer_class = WebsiteFAQSerializer
    permission_classes = [AllowAny]


class WebsitePartnerListCreateView(generics.ListCreateAPIView):
    """Get all partners or create a new one"""
    queryset = WebsitePartner.objects.filter(is_active=True).order_by('order')
    serializer_class = WebsitePartnerSerializer
    permission_classes = [AllowAny]


class WebsitePartnerDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Get, update, or delete a partner"""
    queryset = WebsitePartner.objects.all()
    serializer_class = WebsitePartnerSerializer
    permission_classes = [AllowAny]


class WebsiteHeroSectionView(generics.RetrieveUpdateAPIView):
    """Get or update hero section"""
    queryset = WebsiteHeroSection.objects.all()
    serializer_class = WebsiteHeroSectionSerializer
    permission_classes = [AllowAny]

    def get_object(self):
        obj, created = WebsiteHeroSection.objects.get_or_create(pk=1)
        return obj


class WebsiteCollectionsSectionView(generics.RetrieveUpdateAPIView):
    """Get or update collections section"""
    queryset = WebsiteCollectionsSection.objects.all()
    serializer_class = WebsiteCollectionsSectionSerializer
    permission_classes = [AllowAny]

    def get_object(self):
        obj, created = WebsiteCollectionsSection.objects.get_or_create(pk=1)
        return obj


class WebsiteQualitySectionView(generics.RetrieveUpdateAPIView):
    """Get or update quality/about section"""
    queryset = WebsiteQualitySection.objects.all()
    serializer_class = WebsiteQualitySectionSerializer
    permission_classes = [AllowAny]

    def get_object(self):
        obj, created = WebsiteQualitySection.objects.get_or_create(pk=1)
        return obj


class WebsiteNewsletterView(generics.RetrieveUpdateAPIView):
    """Get or update newsletter settings"""
    queryset = WebsiteNewsletter.objects.all()
    serializer_class = WebsiteNewsletterSerializer
    permission_classes = [AllowAny]

    def get_object(self):
        obj, created = WebsiteNewsletter.objects.get_or_create(pk=1)
        return obj


class WebsiteFurnitureDetailsSectionView(generics.RetrieveUpdateAPIView):
    """Get or update furniture details section"""
    queryset = WebsiteFurnitureDetailsSection.objects.all()
    serializer_class = WebsiteFurnitureDetailsSectionSerializer
    permission_classes = [AllowAny]

    def get_object(self):
        obj, created = WebsiteFurnitureDetailsSection.objects.get_or_create(pk=1)
        return obj


class WebsiteTestimonialsSectionSettingsView(generics.RetrieveUpdateAPIView):
    """Get or update testimonials section settings"""
    queryset = WebsiteTestimonialsSectionSettings.objects.all()
    serializer_class = WebsiteTestimonialsSectionSettingsSerializer
    permission_classes = [AllowAny]

    def get_object(self):
        obj, created = WebsiteTestimonialsSectionSettings.objects.get_or_create(pk=1)
        return obj


class WebsiteStoriesSectionSettingsView(generics.RetrieveUpdateAPIView):
    """Get or update stories section settings"""
    queryset = WebsiteStoriesSectionSettings.objects.all()
    serializer_class = WebsiteStoriesSectionSettingsSerializer
    permission_classes = [AllowAny]

    def get_object(self):
        obj, created = WebsiteStoriesSectionSettings.objects.get_or_create(pk=1)
        return obj


@method_decorator(csrf_exempt, name='dispatch')
class WebsiteDataBulkSaveView(APIView):
    """Save all website data at once (from Website Controller)"""
    permission_classes = [AllowAny]
    
    def post(self, request):
        try:
            data = request.data
            
            # Save stories
            if 'stories' in data:
                for story in data['stories']:
                    # Check if this is a frontend-generated ID (large number from Date.now()) or actual DB ID
                    story_id = story.get('id')
                    image_data = story.get('image_url') or story.get('image', '')
                    if story_id and isinstance(story_id, (int, float)) and story_id > 1000000:
                        # This is a frontend ID, create new record
                        WebsiteStory.objects.create(
                            title=story.get('title', ''),
                            excerpt=story.get('excerpt', ''),
                            author=story.get('author', ''),
                            image_url=image_data,
                        )
                    elif story_id and WebsiteStory.objects.filter(id=story_id).exists():
                        # This is an actual DB ID, update existing
                        WebsiteStory.objects.filter(id=story_id).update(
                            title=story.get('title', ''),
                            excerpt=story.get('excerpt', ''),
                            author=story.get('author', ''),
                            image_url=image_data,
                        )
                    else:
                        # No ID or unknown ID, create new
                        WebsiteStory.objects.create(
                            title=story.get('title', ''),
                            excerpt=story.get('excerpt', ''),
                            author=story.get('author', ''),
                            image_url=image_data,
                        )
            
            # Save testimonials
            if 'testimonials' in data:
                for testimonial in data['testimonials']:
                    testimonial_id = testimonial.get('id')
                    image_data = testimonial.get('image_url') or testimonial.get('image', '')
                    if testimonial_id and isinstance(testimonial_id, (int, float)) and testimonial_id > 1000000:
                        # Frontend ID, create new
                        WebsiteTestimonial.objects.create(
                            name=testimonial.get('name', ''),
                            role=testimonial.get('role', ''),
                            comment=testimonial.get('comment', ''),
                            rating=testimonial.get('rating', 5),
                            image_url=image_data,
                        )
                    elif testimonial_id and WebsiteTestimonial.objects.filter(id=testimonial_id).exists():
                        # DB ID, update existing
                        WebsiteTestimonial.objects.filter(id=testimonial_id).update(
                            name=testimonial.get('name', ''),
                            role=testimonial.get('role', ''),
                            comment=testimonial.get('comment', ''),
                            rating=testimonial.get('rating', 5),
                            image_url=image_data,
                        )
                    else:
                        # No ID, create new
                        WebsiteTestimonial.objects.create(
                            name=testimonial.get('name', ''),
                            role=testimonial.get('role', ''),
                            comment=testimonial.get('comment', ''),
                            rating=testimonial.get('rating', 5),
                            image_url=image_data,
                        )
            
            # Save gallery
            if 'gallery' in data:
                for item in data['gallery']:
                    item_id = item.get('id')
                    image_data = item.get('image_url') or item.get('image', '')
                    if item_id and isinstance(item_id, (int, float)) and item_id > 1000000:
                        # Frontend ID, create new
                        WebsiteGallery.objects.create(
                            title=item.get('title', ''),
                            image_url=image_data,
                            category=item.get('category', 'rooms'),
                        )
                    elif item_id and WebsiteGallery.objects.filter(id=item_id).exists():
                        # DB ID, update existing
                        WebsiteGallery.objects.filter(id=item_id).update(
                            title=item.get('title', ''),
                            image_url=image_data,
                            category=item.get('category', 'rooms'),
                        )
                    else:
                        # No ID, create new
                        WebsiteGallery.objects.create(
                            title=item.get('title', ''),
                            image_url=image_data,
                            category=item.get('category', 'rooms'),
                        )
            
            # Save FAQs
            if 'faqs' in data:
                for faq in data['faqs']:
                    faq_id = faq.get('id')
                    if faq_id and isinstance(faq_id, (int, float)) and faq_id > 1000000:
                        # Frontend ID, create new
                        WebsiteFAQ.objects.create(
                            question=faq.get('question', ''),
                            answer=faq.get('answer', ''),
                            category=faq.get('category', 'shipping'),
                        )
                    elif faq_id and WebsiteFAQ.objects.filter(id=faq_id).exists():
                        # DB ID, update existing
                        WebsiteFAQ.objects.filter(id=faq_id).update(
                            question=faq.get('question', ''),
                            answer=faq.get('answer', ''),
                            category=faq.get('category', 'shipping'),
                        )
                    else:
                        # No ID, create new
                        WebsiteFAQ.objects.create(
                            question=faq.get('question', ''),
                            answer=faq.get('answer', ''),
                            category=faq.get('category', 'shipping'),
                        )
            
            # Save partners
            if 'partners' in data:
                for partner in data['partners']:
                    partner_id = partner.get('id')
                    if partner_id and isinstance(partner_id, (int, float)) and partner_id > 1000000:
                        # Frontend ID, create new
                        WebsitePartner.objects.create(
                            name=partner.get('name', ''),
                            logo_url=partner.get('logo', ''),
                            link=partner.get('link', ''),
                        )
                    elif partner_id and WebsitePartner.objects.filter(id=partner_id).exists():
                        # DB ID, update existing
                        WebsitePartner.objects.filter(id=partner_id).update(
                            name=partner.get('name', ''),
                            logo_url=partner.get('logo', ''),
                            link=partner.get('link', ''),
                        )
                    else:
                        # No ID, create new
                        WebsitePartner.objects.create(
                            name=partner.get('name', ''),
                            logo_url=partner.get('logo', ''),
                            link=partner.get('link', ''),
                        )
            
            # Save hero section
            if 'heroSection' in data:
                hero_data = data['heroSection']
                hero, created = WebsiteHeroSection.objects.get_or_create(pk=1)
                hero.title = hero_data.get('title', '')
                hero.subtitle = hero_data.get('subtitle', '')
                hero.cta_text = hero_data.get('cta_text', 'Explore Products')
                hero.image_url = hero_data.get('image', '')
                hero.save()
            
            # Save newsletter
            if 'newsletter' in data:
                newsletter_data = data['newsletter']
                newsletter, created = WebsiteNewsletter.objects.get_or_create(pk=1)
                newsletter.title = newsletter_data.get('title', 'Subscribe to Our Newsletter')
                newsletter.description = newsletter_data.get('description', '')
                newsletter.placeholder = newsletter_data.get('placeholder', 'Enter your email')
                newsletter.save()
            
            return Response({
                'message': 'All website data saved successfully to database'
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# =============== IMAGE UPLOAD VIEW ===============
class ImageUploadView(APIView):
    """Upload image file and return URL"""
    permission_classes = [AllowAny]
    
    def post(self, request):
        try:
            if 'image' not in request.FILES:
                return Response({'error': 'No image file provided'}, status=status.HTTP_400_BAD_REQUEST)
            
            image_file = request.FILES['image']
            
            # Validate file type
            allowed_extensions = ['jpg', 'jpeg', 'png', 'gif', 'webp']
            file_ext = image_file.name.split('.')[-1].lower()
            
            if file_ext not in allowed_extensions:
                return Response({'error': f'File type {file_ext} not allowed. Use: {", ".join(allowed_extensions)}'}, 
                              status=status.HTTP_400_BAD_REQUEST)
            
            # Create media directory if it doesn't exist
            import os
            from django.conf import settings
            media_dir = settings.MEDIA_ROOT
            if not os.path.exists(media_dir):
                os.makedirs(media_dir)
            
            # Generate unique filename
            from django.utils.text import slugify
            import uuid
            filename = f"{slugify(image_file.name.split('.')[0])}-{uuid.uuid4().hex[:8]}.{file_ext}"
            filepath = os.path.join(media_dir, filename)
            
            # Save file
            with open(filepath, 'wb+') as f:
                for chunk in image_file.chunks():
                    f.write(chunk)
            
            # Return URL
            from django.urls import reverse
            image_url = f"http://localhost:8000/media/{filename}"
            
            return Response({
                'success': True,
                'image_url': image_url,
                'message': 'Image uploaded successfully'
            }, status=status.HTTP_201_CREATED)
        
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)